"""
IBom Selector - Application de sélection de composants PCB
Version 2.0 - Mise à jour avec rendu fidèle à InteractiveHtmlBom

Fonctionnalités:
- Rendu PCB fidèle à IBom (pads, tracks, silkscreen, edges)
- Thème sombre/clair
- Barre de progression
- Navigation Précédent/Suivant
- Options configurables (groupement, taille police, silkscreen)
- Historique des sélections
- Export Excel/CSV
"""

import csv
import json
import re
import math
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("Installation de openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

try:
    from lzstring import LZString as LZStringLib
    HAS_LZSTRING = True
except ImportError:
    HAS_LZSTRING = False
    print("lzstring non disponible, utilisation du décompresseur intégré")


# ==================== THEMES ====================

THEMES = {
    'dark': {
        'bg_primary': '#1a1a2e',
        'bg_secondary': '#16213e',
        'bg_tertiary': '#0f3460',
        'text_primary': '#ffffff',
        'text_secondary': '#b0b0b0',
        'accent': '#e94560',
        'success': '#4ecca3',
        'warning': '#ffc107',
        'border': '#4a4a6a',
        'pcb_board': '#1a1a2e',
        'pcb_edge': '#ffcc00',
        'pad_front': '#b0a050',
        'pad_back': '#5050a0',
        'pad_hole': '#cccccc',
        'pad_highlight': '#d04040',
        'track_front': '#def5f1',
        'track_back': '#42524f',
        'silk_edge': '#aaaa44',
        'silk_text': '#44aaaa',
        'selection_rect': '#ffcc00',
        'progress_bg': '#2a2a4a',
        'progress_fill': '#4ecca3',
        'row_done': '#1a3a2a',
        'row_pending': '#1a1a2e',
    },
    'light': {
        'bg_primary': '#ffffff',
        'bg_secondary': '#f5f5f5',
        'bg_tertiary': '#e0e0e0',
        'text_primary': '#1a1a1a',
        'text_secondary': '#666666',
        'accent': '#e94560',
        'success': '#28a745',
        'warning': '#ffc107',
        'border': '#cccccc',
        'pcb_board': '#2a4a2a',
        'pcb_edge': '#ffcc00',
        'pad_front': '#c0b060',
        'pad_back': '#6060b0',
        'pad_hole': '#ffffff',
        'pad_highlight': '#ff4444',
        'track_front': '#90c090',
        'track_back': '#607060',
        'silk_edge': '#cccc66',
        'silk_text': '#66cccc',
        'selection_rect': '#ff6600',
        'progress_bg': '#e0e0e0',
        'progress_fill': '#28a745',
        'row_done': '#c8e6c9',
        'row_pending': '#ffffff',
    }
}


# ==================== PREFERENCES ====================

class Preferences:
    """Gestionnaire de préférences persistées"""
    
    DEFAULT = {
        'theme': 'dark',
        'font_size': 11,
        'group_by_value': True,
        'show_silkscreen': True,
        'show_tracks': True,
        'show_pads': True,
        'auto_save': False,
        'auto_save_minutes': 5,
    }
    
    def __init__(self):
        self.prefs = self.DEFAULT.copy()
        self.file_path = Path.home() / '.ibom_selector_prefs.json'
        self.load()
    
    def load(self):
        """Charge les préférences depuis le fichier"""
        try:
            if self.file_path.exists():
                with open(self.file_path, 'r') as f:
                    saved = json.load(f)
                    self.prefs.update(saved)
        except Exception as e:
            print(f"Erreur chargement préférences: {e}")
    
    def save(self):
        """Sauvegarde les préférences"""
        try:
            with open(self.file_path, 'w') as f:
                json.dump(self.prefs, f, indent=2)
        except Exception as e:
            print(f"Erreur sauvegarde préférences: {e}")
    
    def get(self, key, default=None):
        return self.prefs.get(key, default)
    
    def set(self, key, value):
        self.prefs[key] = value
        self.save()


# ==================== LZ-STRING DECOMPRESSOR ====================

class LZString:
    """Décompresseur LZ-String pour les données InteractiveHtmlBom"""
    
    @staticmethod
    def decompress_from_base64(compressed):
        """Décompresse une chaîne encodée en base64"""
        if not compressed:
            return ""
        
        key_str = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/="
        base_reverse_dict = {char: i for i, char in enumerate(key_str)}
        
        try:
            length = len(compressed)
            get_next_value = lambda index: base_reverse_dict.get(compressed[index], 0)
            result = LZString._decompress(length, 32, get_next_value)
            return result
        except Exception as e:
            print(f"Erreur de décompression: {e}")
            return None
    
    @staticmethod
    def _decompress(length, reset_value, get_next_value):
        """Algorithme de décompression LZ"""
        dictionary = {}
        enlargeIn = 4
        dictSize = 4
        numBits = 3
        entry = ""
        result = []
        
        data_val = get_next_value(0)
        data_position = reset_value
        data_index = 1
        
        for i in range(3):
            dictionary[i] = i
        
        bits = 0
        maxpower = 2 ** 2
        power = 1
        
        while power != maxpower:
            resb = data_val & data_position
            data_position >>= 1
            if data_position == 0:
                data_position = reset_value
                data_val = get_next_value(data_index)
                data_index += 1
            bits |= (1 if resb > 0 else 0) * power
            power <<= 1
        
        next_val = bits
        if next_val == 0:
            bits = 0
            maxpower = 2 ** 8
            power = 1
            while power != maxpower:
                resb = data_val & data_position
                data_position >>= 1
                if data_position == 0:
                    data_position = reset_value
                    data_val = get_next_value(data_index)
                    data_index += 1
                bits |= (1 if resb > 0 else 0) * power
                power <<= 1
            c = chr(bits)
        elif next_val == 1:
            bits = 0
            maxpower = 2 ** 16
            power = 1
            while power != maxpower:
                resb = data_val & data_position
                data_position >>= 1
                if data_position == 0:
                    data_position = reset_value
                    data_val = get_next_value(data_index)
                    data_index += 1
                bits |= (1 if resb > 0 else 0) * power
                power <<= 1
            c = chr(bits)
        elif next_val == 2:
            return ""
        
        dictionary[3] = c
        w = c
        result.append(c)
        
        while True:
            if data_index > length:
                return ""
            
            bits = 0
            maxpower = 2 ** numBits
            power = 1
            while power != maxpower:
                resb = data_val & data_position
                data_position >>= 1
                if data_position == 0:
                    data_position = reset_value
                    data_val = get_next_value(data_index)
                    data_index += 1
                bits |= (1 if resb > 0 else 0) * power
                power <<= 1
            
            c = bits
            if c == 0:
                bits = 0
                maxpower = 2 ** 8
                power = 1
                while power != maxpower:
                    resb = data_val & data_position
                    data_position >>= 1
                    if data_position == 0:
                        data_position = reset_value
                        data_val = get_next_value(data_index)
                        data_index += 1
                    bits |= (1 if resb > 0 else 0) * power
                    power <<= 1
                dictionary[dictSize] = chr(bits)
                dictSize += 1
                c = dictSize - 1
                enlargeIn -= 1
            elif c == 1:
                bits = 0
                maxpower = 2 ** 16
                power = 1
                while power != maxpower:
                    resb = data_val & data_position
                    data_position >>= 1
                    if data_position == 0:
                        data_position = reset_value
                        data_val = get_next_value(data_index)
                        data_index += 1
                    bits |= (1 if resb > 0 else 0) * power
                    power <<= 1
                dictionary[dictSize] = chr(bits)
                dictSize += 1
                c = dictSize - 1
                enlargeIn -= 1
            elif c == 2:
                return "".join(result)
            
            if enlargeIn == 0:
                enlargeIn = 2 ** numBits
                numBits += 1
            
            if c in dictionary:
                entry = dictionary[c] if isinstance(dictionary[c], str) else chr(dictionary[c])
            else:
                if c == dictSize:
                    entry = w + w[0]
                else:
                    return None
            
            result.append(entry)
            dictionary[dictSize] = w + entry[0]
            dictSize += 1
            enlargeIn -= 1
            
            if enlargeIn == 0:
                enlargeIn = 2 ** numBits
                numBits += 1
            
            w = entry
        
        return "".join(result)


# ==================== IBOM PARSER ====================

class IBomParser:
    """Parse le fichier HTML d'InteractiveHtmlBom pour extraire les données"""
    
    def __init__(self, html_file_path):
        self.html_file_path = html_file_path
        self.pcbdata = None
        self.components = []
        self.bom_data = []
        self.board_bbox = None
        self.lcsc_data = {}
        self.footprints = []
        self.edges = []
        self.tracks = {}
        self.drawings = {}
        
    def _load_lcsc_csv(self):
        """Charge le fichier CSV LCSC s'il existe"""
        html_dir = Path(self.html_file_path).parent
        possible_paths = [
            html_dir.parent / 'lcsc' / 'BOM-lcsc.csv',
            html_dir / 'lcsc' / 'BOM-lcsc.csv',
            Path('lcsc') / 'BOM-lcsc.csv',
            html_dir.parent / 'BOM-lcsc.csv',
        ]
        
        csv_path = None
        for path in possible_paths:
            if path.exists():
                csv_path = path
                break
        
        if not csv_path:
            return
        
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    designators = row.get('Designator', '')
                    lcsc_code = row.get('LCSC', '').strip()
                    
                    if lcsc_code:
                        for ref in designators.split(','):
                            ref = ref.strip()
                            if ref:
                                self.lcsc_data[ref] = lcsc_code
            
            print(f"Fichier LCSC chargé: {len(self.lcsc_data)} références")
        except Exception as e:
            print(f"Erreur lors du chargement du fichier LCSC: {e}")
        
    def parse(self):
        """Parse le fichier HTML et extrait les données PCB"""
        with open(self.html_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Chercher les données compressées
        lz_match = re.search(r'LZString\.decompressFromBase64\(["\']([^"\']+)["\']\)', content)
        
        if lz_match:
            compressed_data = lz_match.group(1)
            print(f"Données compressées trouvées ({len(compressed_data)} caractères)")
            
            try:
                if HAS_LZSTRING:
                    lz = LZStringLib()
                    decompressed = lz.decompressFromBase64(compressed_data)
                else:
                    decompressed = LZString.decompress_from_base64(compressed_data)
                
                if decompressed:
                    self.pcbdata = json.loads(decompressed)
                    print(f"Décompression réussie!")
                else:
                    raise ValueError("Échec de la décompression LZ-String")
            except Exception as e:
                print(f"Erreur lors de la décompression: {e}")
                raise ValueError(f"Impossible de décompresser les données: {e}")
        else:
            pcbdata_match = re.search(r'var\s+pcbdata\s*=\s*(\{.*?\});', content, re.DOTALL)
            if not pcbdata_match:
                pcbdata_match = re.search(r'pcbdata\s*=\s*(\{.*?\})\s*[;\n]', content, re.DOTALL)
            
            if pcbdata_match:
                pcbdata_str = pcbdata_match.group(1)
                try:
                    self.pcbdata = json.loads(pcbdata_str)
                except json.JSONDecodeError:
                    pcbdata_str = re.sub(r',\s*}', '}', pcbdata_str)
                    pcbdata_str = re.sub(r',\s*]', ']', pcbdata_str)
                    self.pcbdata = json.loads(pcbdata_str)
            else:
                raise ValueError("Impossible de trouver les données pcbdata dans le fichier HTML")
        
        self._load_lcsc_csv()
        self._extract_footprints()
        self._extract_components()
        self._extract_bom()
        self._extract_edges()
        self._extract_tracks()
        self._extract_drawings()
        self._calculate_board_bbox()
        
        return self
    
    def _extract_footprints(self):
        """Extrait les footprints avec leurs pads et drawings"""
        self.footprints = self.pcbdata.get('footprints', [])
        print(f"Footprints extraits: {len(self.footprints)}")
    
    def _extract_edges(self):
        """Extrait les edges (contour du PCB)"""
        self.edges = self.pcbdata.get('edges', [])
        print(f"Edges extraits: {len(self.edges)}")
    
    def _extract_tracks(self):
        """Extrait les tracks (pistes de cuivre)"""
        self.tracks = self.pcbdata.get('tracks', {})
        total = sum(len(t) for t in self.tracks.values() if isinstance(t, list))
        print(f"Tracks extraits: {total}")
    
    def _extract_drawings(self):
        """Extrait les drawings (silkscreen, etc.)"""
        self.drawings = self.pcbdata.get('drawings', {})
    
    def _extract_components(self):
        """Extrait les composants avec leurs positions"""
        self.components = []
        
        for fp_id, fp in enumerate(self.footprints):
            ref = fp.get('ref', '')
            layer = fp.get('layer', 'F')
            bbox = fp.get('bbox', {})
            
            x, y = 0, 0
            if bbox and 'pos' in bbox:
                pos = bbox.get('pos', [0, 0])
                if isinstance(pos, list) and len(pos) >= 2:
                    x, y = pos[0], pos[1]
            elif 'center' in fp:
                center = fp.get('center', [0, 0])
                if isinstance(center, list) and len(center) >= 2:
                    x, y = center[0], center[1]
            elif bbox and 'minx' in bbox:
                x = (bbox.get('minx', 0) + bbox.get('maxx', 0)) / 2
                y = (bbox.get('miny', 0) + bbox.get('maxy', 0)) / 2
            else:
                pads = fp.get('pads', [])
                if pads:
                    x = sum(p.get('pos', [0, 0])[0] for p in pads) / len(pads)
                    y = sum(p.get('pos', [0, 0])[1] for p in pads) / len(pads)
            
            self.components.append({
                'ref': ref,
                'id': fp_id,
                'x': x,
                'y': y,
                'layer': layer,
                'bbox': bbox
            })
    
    def _extract_bom(self):
        """Extrait les données BOM"""
        self.bom_data = []
        
        bom = self.pcbdata.get('bom', {})
        fields_data = bom.get('fields', {})
        both = bom.get('both', [])
        
        for group in both:
            if not isinstance(group, list):
                continue
                
            for ref_item in group:
                if not isinstance(ref_item, list) or len(ref_item) < 2:
                    continue
                    
                ref_name = ref_item[0]
                fp_id = ref_item[1]
                
                value = ''
                footprint_name = ''
                lcsc = ''
                
                fp_id_str = str(fp_id)
                if fp_id_str in fields_data:
                    component_fields = fields_data[fp_id_str]
                    if len(component_fields) >= 1:
                        value = component_fields[0] or ''
                    if len(component_fields) >= 2:
                        footprint_name = component_fields[1] or ''
                    for field_val in component_fields:
                        if isinstance(field_val, str) and len(field_val) > 1:
                            if field_val.startswith('C') and field_val[1:].isdigit():
                                lcsc = field_val
                                break
                
                if not lcsc and ref_name in self.lcsc_data:
                    lcsc = self.lcsc_data[ref_name]
                
                self.bom_data.append({
                    'ref': ref_name,
                    'id': fp_id,
                    'value': value,
                    'footprint': footprint_name,
                    'lcsc': lcsc
                })
    
    def _calculate_board_bbox(self):
        """Calcule la bounding box du PCB"""
        edges = self.pcbdata.get('edges_bbox', {})
        if edges:
            self.board_bbox = {
                'minx': edges.get('minx', 0),
                'miny': edges.get('miny', 0),
                'maxx': edges.get('maxx', 100),
                'maxy': edges.get('maxy', 100)
            }
        else:
            if self.components:
                xs = [c['x'] for c in self.components]
                ys = [c['y'] for c in self.components]
                margin = 5
                self.board_bbox = {
                    'minx': min(xs) - margin,
                    'miny': min(ys) - margin,
                    'maxx': max(xs) + margin,
                    'maxy': max(ys) + margin
                }
            else:
                self.board_bbox = {'minx': 0, 'miny': 0, 'maxx': 100, 'maxy': 100}
    
    def get_bom_for_ref(self, ref, fp_id=None):
        """Récupère les infos BOM pour une référence"""
        for bom_entry in self.bom_data:
            if bom_entry['ref'] == ref:
                return bom_entry
            if fp_id is not None and bom_entry.get('id') == fp_id:
                return bom_entry
        return {'ref': ref, 'value': '', 'footprint': '', 'lcsc': '', 'id': fp_id}
    
    def get_components_in_rect(self, x1, y1, x2, y2):
        """Retourne les composants dans le rectangle spécifié"""
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
        
        selected = []
        for comp in self.components:
            if min_x <= comp['x'] <= max_x and min_y <= comp['y'] <= max_y:
                bom_info = self.get_bom_for_ref(comp['ref'], comp.get('id'))
                selected.append({
                    'ref': comp['ref'],
                    'value': bom_info.get('value', ''),
                    'footprint': bom_info.get('footprint', ''),
                    'lcsc': bom_info.get('lcsc', ''),
                    'x': comp['x'],
                    'y': comp['y'],
                    'layer': comp['layer']
                })
        
        return selected


# ==================== PCB VIEWER ====================

class PCBViewer(tk.Toplevel):
    """Fenêtre de visualisation du PCB avec rendu fidèle à IBom"""
    
    def __init__(self, parent, parser, callback, prefs, theme):
        super().__init__(parent)
        self.parser = parser
        self.callback = callback
        self.prefs = prefs
        self.theme = theme
        
        self.title("Sélection de zone - PCB Viewer")
        self.geometry("1000x800")
        self.configure(bg=theme['bg_primary'])
        
        # Variables de sélection
        self.start_x = None
        self.start_y = None
        self.rect_id = None
        self.scale = 1.0
        self.offset_x = 50
        self.offset_y = 50
        self.pan_start_x = None
        self.pan_start_y = None
        
        self._setup_ui()
        self.after(100, self._draw_pcb)
    
    def _setup_ui(self):
        """Configure l'interface utilisateur"""
        main_frame = tk.Frame(self, bg=self.theme['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Instructions
        instructions = tk.Label(
            main_frame, 
            text="Clic gauche + glisser = Sélection | Molette = Zoom | Clic droit + glisser = Pan",
            font=('Segoe UI', 10),
            bg=self.theme['bg_primary'],
            fg=self.theme['text_secondary']
        )
        instructions.pack(pady=(0, 10))
        
        # Options frame
        options_frame = tk.Frame(main_frame, bg=self.theme['bg_primary'])
        options_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Checkboxes pour les layers
        self.show_pads_var = tk.BooleanVar(value=self.prefs.get('show_pads', True))
        self.show_tracks_var = tk.BooleanVar(value=self.prefs.get('show_tracks', True))
        self.show_silk_var = tk.BooleanVar(value=self.prefs.get('show_silkscreen', True))
        
        tk.Checkbutton(options_frame, text="Pads", variable=self.show_pads_var,
                       command=self._draw_pcb, bg=self.theme['bg_primary'], 
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_secondary']
                       ).pack(side=tk.LEFT, padx=10)
        tk.Checkbutton(options_frame, text="Tracks", variable=self.show_tracks_var,
                       command=self._draw_pcb, bg=self.theme['bg_primary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_secondary']
                       ).pack(side=tk.LEFT, padx=10)
        tk.Checkbutton(options_frame, text="Silkscreen", variable=self.show_silk_var,
                       command=self._draw_pcb, bg=self.theme['bg_primary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_secondary']
                       ).pack(side=tk.LEFT, padx=10)
        
        # Canvas pour le PCB
        canvas_frame = tk.Frame(main_frame, bg=self.theme['bg_primary'])
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas = tk.Canvas(canvas_frame, bg=self.theme['pcb_board'], highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Bindings
        self.canvas.bind('<Button-1>', self._on_mouse_down)
        self.canvas.bind('<B1-Motion>', self._on_mouse_drag)
        self.canvas.bind('<ButtonRelease-1>', self._on_mouse_up)
        self.canvas.bind('<Button-3>', self._on_pan_start)
        self.canvas.bind('<B3-Motion>', self._on_pan_drag)
        self.canvas.bind('<MouseWheel>', self._on_mousewheel)
        self.canvas.bind('<Button-4>', lambda e: self._zoom_in())  # Linux
        self.canvas.bind('<Button-5>', lambda e: self._zoom_out())  # Linux
        
        # Boutons
        btn_frame = tk.Frame(main_frame, bg=self.theme['bg_primary'])
        btn_frame.pack(pady=10)
        
        btn_style = {'bg': self.theme['bg_tertiary'], 'fg': self.theme['text_primary'],
                     'activebackground': self.theme['accent'], 'activeforeground': '#ffffff',
                     'relief': tk.FLAT, 'padx': 15, 'pady': 5}
        
        tk.Button(btn_frame, text="Zoom +", command=self._zoom_in, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Zoom -", command=self._zoom_out, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Réinitialiser", command=self._reset_view, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Annuler", command=self.destroy, **btn_style).pack(side=tk.LEFT, padx=5)
        
        # Légende
        legend_frame = tk.Frame(main_frame, bg=self.theme['bg_primary'])
        legend_frame.pack(pady=5)
        
        tk.Label(legend_frame, text="● Front", fg=self.theme['pad_front'], 
                 bg=self.theme['bg_primary'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=10)
        tk.Label(legend_frame, text="● Back", fg=self.theme['pad_back'],
                 bg=self.theme['bg_primary'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=10)
        tk.Label(legend_frame, text="● Highlight", fg=self.theme['pad_highlight'],
                 bg=self.theme['bg_primary'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=10)
    
    def _pcb_to_canvas(self, x, y):
        """Convertit les coordonnées PCB en coordonnées canvas (Y inversé)"""
        bbox = self.parser.board_bbox
        canvas_x = self.offset_x + (x - bbox['minx']) * self.scale
        # Inverser Y pour avoir le même sens que IBom
        canvas_height = self.canvas.winfo_height() or 700
        canvas_y = canvas_height - (self.offset_y + (y - bbox['miny']) * self.scale)
        return canvas_x, canvas_y
    
    def _canvas_to_pcb(self, canvas_x, canvas_y):
        """Convertit les coordonnées canvas en coordonnées PCB"""
        bbox = self.parser.board_bbox
        canvas_height = self.canvas.winfo_height() or 700
        x = (canvas_x - self.offset_x) / self.scale + bbox['minx']
        y = (canvas_height - canvas_y - self.offset_y) / self.scale + bbox['miny']
        return x, y
    
    def _draw_pcb(self, recalculate_scale=True):
        """Dessine le PCB avec tous les éléments"""
        self.canvas.delete('all')
        
        bbox = self.parser.board_bbox
        width = bbox['maxx'] - bbox['minx']
        height = bbox['maxy'] - bbox['miny']
        
        canvas_width = self.canvas.winfo_width() or 900
        canvas_height = self.canvas.winfo_height() or 700
        
        if recalculate_scale:
            scale_x = (canvas_width - 100) / width if width > 0 else 1
            scale_y = (canvas_height - 100) / height if height > 0 else 1
            self.scale = min(scale_x, scale_y) * 0.9
            
            # Centrer
            self.offset_x = (canvas_width - width * self.scale) / 2
            self.offset_y = (canvas_height - height * self.scale) / 2
        
        # Fond du PCB
        x1, y1 = self._pcb_to_canvas(bbox['minx'], bbox['miny'])
        x2, y2 = self._pcb_to_canvas(bbox['maxx'], bbox['maxy'])
        self.canvas.create_rectangle(min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2), 
                                      outline=self.theme['pcb_edge'], fill=self.theme['pcb_board'], width=2)
        
        # Dessiner dans l'ordre
        self._draw_edges()
        if self.show_tracks_var.get():
            self._draw_tracks()
        if self.show_pads_var.get():
            self._draw_pads()
        if self.show_silk_var.get():
            self._draw_silkscreen()
    
    def _draw_edges(self):
        """Dessine le contour du PCB"""
        for edge in self.parser.edges:
            edge_type = edge.get('type', '')
            
            if edge_type == 'segment':
                start = edge.get('start', [0, 0])
                end = edge.get('end', [0, 0])
                x1, y1 = self._pcb_to_canvas(start[0], start[1])
                x2, y2 = self._pcb_to_canvas(end[0], end[1])
                width = max(1, edge.get('width', 0.15) * self.scale)
                self.canvas.create_line(x1, y1, x2, y2, fill=self.theme['pcb_edge'], width=width)
            
            elif edge_type == 'circle':
                center = edge.get('start', [0, 0])
                radius = edge.get('radius', 1) * self.scale
                cx, cy = self._pcb_to_canvas(center[0], center[1])
                self.canvas.create_oval(cx - radius, cy - radius, cx + radius, cy + radius,
                                        outline=self.theme['pcb_edge'], width=1)
            
            elif edge_type == 'arc':
                start = edge.get('start', [0, 0])
                radius = edge.get('radius', 1)
                start_angle = edge.get('startangle', 0)
                end_angle = edge.get('endangle', 360)
                cx, cy = self._pcb_to_canvas(start[0], start[1])
                r = radius * self.scale
                
                # Dessiner l'arc comme une série de segments
                num_segments = 20
                points = []
                for i in range(num_segments + 1):
                    angle = math.radians(start_angle + (end_angle - start_angle) * i / num_segments)
                    px = cx + r * math.cos(angle)
                    py = cy - r * math.sin(angle)  # Négatif car Y inversé
                    points.extend([px, py])
                
                if len(points) >= 4:
                    self.canvas.create_line(points, fill=self.theme['pcb_edge'], width=1, smooth=True)
    
    def _draw_tracks(self):
        """Dessine les pistes de cuivre"""
        for layer, layer_tracks in self.parser.tracks.items():
            if not isinstance(layer_tracks, list):
                continue
            
            is_front = layer.startswith('F') or layer == 'F.Cu'
            color = self.theme['track_front'] if is_front else self.theme['track_back']
            
            for track in layer_tracks:
                start = track.get('start')
                end = track.get('end')
                width = track.get('width', 0.2)
                
                if start and end:
                    x1, y1 = self._pcb_to_canvas(start[0], start[1])
                    x2, y2 = self._pcb_to_canvas(end[0], end[1])
                    stroke_width = max(0.8, width * self.scale)
                    self.canvas.create_line(x1, y1, x2, y2, fill=color, width=stroke_width, 
                                           capstyle=tk.ROUND)
    
    def _draw_pads(self):
        """Dessine les pads de tous les footprints"""
        for fp in self.parser.footprints:
            fp_layer = fp.get('layer', 'F')
            pads = fp.get('pads', [])
            
            for pad in pads:
                self._draw_pad(pad, fp_layer)
    
    def _draw_pad(self, pad, fp_layer):
        """Dessine un pad individuel avec sa forme exacte"""
        pos = pad.get('pos', [0, 0])
        size = pad.get('size', [0.5, 0.5])
        shape = pad.get('shape', 'rect')
        pad_type = pad.get('type', 'smd')
        layers = pad.get('layers', [fp_layer])
        offset = pad.get('offset', [0, 0])
        angle = pad.get('angle', 0)
        radius = pad.get('radius', 0.25)
        drillsize = pad.get('drillsize', [0.3, 0.3])
        drillshape = pad.get('drillshape', 'circle')
        
        # Position avec offset
        actual_x = pos[0] + offset[0]
        actual_y = pos[1] + offset[1]
        cx, cy = self._pcb_to_canvas(actual_x, actual_y)
        
        w = max(2, size[0] * self.scale)
        h = max(2, size[1] * self.scale)
        
        # Couleur selon la couche
        is_front = 'F' in layers or any(l.startswith('F.') for l in layers)
        color = self.theme['pad_front'] if is_front else self.theme['pad_back']
        
        # Dessiner selon la forme
        if shape == 'circle':
            r = w / 2
            self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r, fill=color, outline='')
        elif shape == 'oval':
            # Ellipse
            self.canvas.create_oval(cx - w/2, cy - h/2, cx + w/2, cy + h/2, fill=color, outline='')
        elif shape == 'roundrect':
            # Rectangle arrondi - approximer avec un rectangle
            corner_radius = min(w, h) * radius
            self._draw_rounded_rect(cx - w/2, cy - h/2, cx + w/2, cy + h/2, corner_radius, color)
        else:
            # Rectangle standard
            self.canvas.create_rectangle(cx - w/2, cy - h/2, cx + w/2, cy + h/2, fill=color, outline='')
        
        # Dessiner le trou pour les pads through-hole
        if pad_type == 'th' and drillsize:
            hole_w = max(1.5, drillsize[0] * self.scale)
            hole_h = drillsize[1] * self.scale if len(drillsize) > 1 else hole_w
            
            if drillshape == 'oblong':
                self.canvas.create_oval(cx - hole_w/2, cy - hole_h/2, cx + hole_w/2, cy + hole_h/2,
                                       fill=self.theme['pad_hole'], outline='')
            else:
                r = hole_w / 2
                self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r,
                                       fill=self.theme['pad_hole'], outline='')
    
    def _draw_rounded_rect(self, x1, y1, x2, y2, radius, color):
        """Dessine un rectangle arrondi"""
        points = [
            x1 + radius, y1,
            x2 - radius, y1,
            x2, y1,
            x2, y1 + radius,
            x2, y2 - radius,
            x2, y2,
            x2 - radius, y2,
            x1 + radius, y2,
            x1, y2,
            x1, y2 - radius,
            x1, y1 + radius,
            x1, y1,
        ]
        self.canvas.create_polygon(points, fill=color, outline='', smooth=True)
    
    def _draw_silkscreen(self):
        """Dessine le silkscreen et les références"""
        # Dessiner les drawings des footprints
        for fp in self.parser.footprints:
            drawings = fp.get('drawings', [])
            ref = fp.get('ref', '')
            bbox = fp.get('bbox', {})
            
            for drawing_obj in drawings:
                layer = drawing_obj.get('layer', '')
                drawing = drawing_obj.get('drawing', drawing_obj)
                
                # Seulement silkscreen
                if 'Silk' not in layer and 'SilkS' not in layer:
                    if drawing_obj.get('layer'):
                        continue
                
                self._draw_silkscreen_element(drawing)
            
            # Dessiner la référence du composant
            if ref and bbox:
                self._draw_component_ref(ref, bbox)
    
    def _draw_silkscreen_element(self, drawing):
        """Dessine un élément de silkscreen"""
        draw_type = drawing.get('type', '')
        color = self.theme['silk_edge']
        
        if draw_type == 'segment':
            start = drawing.get('start', [0, 0])
            end = drawing.get('end', [0, 0])
            x1, y1 = self._pcb_to_canvas(start[0], start[1])
            x2, y2 = self._pcb_to_canvas(end[0], end[1])
            width = max(0.5, (drawing.get('width', 0.1)) * self.scale)
            self.canvas.create_line(x1, y1, x2, y2, fill=color, width=width)
        
        elif draw_type == 'rect':
            start = drawing.get('start', [0, 0])
            end = drawing.get('end', [1, 1])
            x1, y1 = self._pcb_to_canvas(start[0], start[1])
            x2, y2 = self._pcb_to_canvas(end[0], end[1])
            width = max(0.5, (drawing.get('width', 0.1)) * self.scale)
            self.canvas.create_rectangle(min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2),
                                        outline=color, width=width)
        
        elif draw_type == 'circle':
            center = drawing.get('start', [0, 0])
            radius = (drawing.get('radius', 0.5)) * self.scale
            cx, cy = self._pcb_to_canvas(center[0], center[1])
            width = max(0.5, (drawing.get('width', 0.1)) * self.scale)
            self.canvas.create_oval(cx - radius, cy - radius, cx + radius, cy + radius,
                                   outline=color, width=width)
        
        elif draw_type == 'polygon':
            polygons = drawing.get('polygons', [])
            for poly in polygons:
                if isinstance(poly, list) and len(poly) >= 3:
                    points = []
                    for pt in poly:
                        if isinstance(pt, list) and len(pt) >= 2:
                            px, py = self._pcb_to_canvas(pt[0], pt[1])
                            points.extend([px, py])
                    if len(points) >= 6:
                        filled = drawing.get('filled', False)
                        if filled:
                            self.canvas.create_polygon(points, fill=color, outline='')
                        else:
                            self.canvas.create_polygon(points, fill='', outline=color, width=1)
    
    def _draw_component_ref(self, ref, bbox):
        """Dessine la référence d'un composant"""
        if not ref or ref == 'REF**':
            return
        
        bbox_pos = bbox.get('pos', [0, 0])
        bbox_relpos = bbox.get('relpos', [0, 0])
        bbox_size = bbox.get('size', [1, 1])
        
        center_x = bbox_pos[0] + bbox_relpos[0] + bbox_size[0] / 2
        center_y = bbox_pos[1] + bbox_relpos[1] + bbox_size[1] / 2
        cx, cy = self._pcb_to_canvas(center_x, center_y)
        
        # Taille de police proportionnelle
        font_size = max(6, min(12, int(min(bbox_size[0], bbox_size[1]) * self.scale * 0.4)))
        
        self.canvas.create_text(cx, cy, text=ref, fill=self.theme['silk_text'],
                               font=('Consolas', font_size, 'bold'))
    
    def _on_mouse_down(self, event):
        """Début de la sélection"""
        self.start_x = event.x
        self.start_y = event.y
        if self.rect_id:
            self.canvas.delete(self.rect_id)
    
    def _on_mouse_drag(self, event):
        """Pendant la sélection"""
        if self.rect_id:
            self.canvas.delete(self.rect_id)
        
        self.rect_id = self.canvas.create_rectangle(
            self.start_x, self.start_y, event.x, event.y,
            outline=self.theme['selection_rect'], width=2, dash=(5, 5)
        )
    
    def _on_mouse_up(self, event):
        """Fin de la sélection"""
        if self.start_x is None:
            return
        
        # Vérifier taille minimale
        if abs(event.x - self.start_x) < 10 or abs(event.y - self.start_y) < 10:
            if self.rect_id:
                self.canvas.delete(self.rect_id)
            return
        
        # Convertir en coordonnées PCB
        pcb_x1, pcb_y1 = self._canvas_to_pcb(self.start_x, self.start_y)
        pcb_x2, pcb_y2 = self._canvas_to_pcb(event.x, event.y)
        
        # Récupérer les composants sélectionnés
        selected = self.parser.get_components_in_rect(pcb_x1, pcb_y1, pcb_x2, pcb_y2)
        
        if selected:
            rect = (min(pcb_x1, pcb_x2), min(pcb_y1, pcb_y2), 
                    max(pcb_x1, pcb_x2), max(pcb_y1, pcb_y2))
            self.callback(selected, rect)
            self.destroy()
        else:
            messagebox.showinfo("Sélection", "Aucun composant dans la zone sélectionnée.\nEssayez une autre zone.")
            if self.rect_id:
                self.canvas.delete(self.rect_id)
    
    def _on_pan_start(self, event):
        """Début du pan"""
        self.pan_start_x = event.x
        self.pan_start_y = event.y
    
    def _on_pan_drag(self, event):
        """Pan en cours"""
        if self.pan_start_x is None:
            return
        
        dx = event.x - self.pan_start_x
        dy = event.y - self.pan_start_y
        
        self.offset_x += dx
        self.offset_y -= dy  # Inverser car Y est inversé
        
        self.pan_start_x = event.x
        self.pan_start_y = event.y
        
        self._draw_pcb(recalculate_scale=False)
    
    def _on_mousewheel(self, event):
        """Zoom avec la molette"""
        if event.delta > 0:
            self._zoom_in()
        else:
            self._zoom_out()
    
    def _zoom_in(self):
        self.scale *= 1.2
        self._draw_pcb(recalculate_scale=False)

    def _zoom_out(self):
        self.scale /= 1.2
        self._draw_pcb(recalculate_scale=False)

    def _reset_view(self):
        self._draw_pcb(recalculate_scale=True)


# ==================== SPLIT VIEW ====================

class SplitView(tk.Toplevel):
    """Fenêtre Split avec PCB et Liste côte à côte, synchronisés"""
    
    def __init__(self, parent, parser, components, processed_items, prefs, theme, on_processed_change):
        super().__init__(parent)
        self.parser = parser
        self.components = components
        self.processed_items = processed_items
        self.prefs = prefs
        self.theme = theme
        self.on_processed_change = on_processed_change
        
        self.title("Split View - PCB + Liste")
        self.geometry("1400x900")
        self.configure(bg=theme['bg_primary'])
        
        # Variables
        self.scale = 1.0
        self.offset_x = 50
        self.offset_y = 50
        self.highlighted_refs = set()
        self.show_pads_var = tk.BooleanVar(value=True)
        self.show_tracks_var = tk.BooleanVar(value=True)
        self.show_silk_var = tk.BooleanVar(value=True)
        self.group_by_value_var = tk.BooleanVar(value=prefs.get('group_by_value', True))
        
        self._setup_ui()
        self.after(100, self._draw_pcb)
    
    def _setup_ui(self):
        """Configure l'interface split"""
        # Frame principal avec PanedWindow
        self.paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg=self.theme['bg_primary'],
                                     sashwidth=5, sashrelief=tk.RAISED)
        self.paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # ===== PANNEAU GAUCHE: PCB =====
        left_frame = tk.Frame(self.paned, bg=self.theme['bg_primary'])
        self.paned.add(left_frame, width=700)
        
        # Toolbar PCB
        pcb_toolbar = tk.Frame(left_frame, bg=self.theme['bg_secondary'])
        pcb_toolbar.pack(fill=tk.X, pady=2)
        
        tk.Label(pcb_toolbar, text="PCB", font=('Segoe UI', 11, 'bold'),
                bg=self.theme['bg_secondary'], fg=self.theme['text_primary']).pack(side=tk.LEFT, padx=10)
        
        tk.Checkbutton(pcb_toolbar, text="Pads", variable=self.show_pads_var,
                       command=self._draw_pcb, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(pcb_toolbar, text="Tracks", variable=self.show_tracks_var,
                       command=self._draw_pcb, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(pcb_toolbar, text="Silk", variable=self.show_silk_var,
                       command=self._draw_pcb, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=5)
        
        btn_style = {'bg': self.theme['bg_tertiary'], 'fg': self.theme['text_primary'],
                     'activebackground': self.theme['accent'], 'activeforeground': '#ffffff',
                     'relief': tk.FLAT, 'padx': 8, 'pady': 2}
        
        tk.Button(pcb_toolbar, text="Zoom +", command=self._zoom_in, **btn_style).pack(side=tk.RIGHT, padx=2)
        tk.Button(pcb_toolbar, text="Zoom -", command=self._zoom_out, **btn_style).pack(side=tk.RIGHT, padx=2)
        tk.Button(pcb_toolbar, text="Reset", command=self._reset_view, **btn_style).pack(side=tk.RIGHT, padx=2)
        
        # Canvas PCB
        self.canvas = tk.Canvas(left_frame, bg=self.theme['pcb_board'], highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Bindings PCB
        self.canvas.bind('<MouseWheel>', self._on_mousewheel)
        self.canvas.bind('<Button-3>', self._on_pan_start)
        self.canvas.bind('<B3-Motion>', self._on_pan_drag)
        self.pan_start_x = None
        self.pan_start_y = None
        
        # Légende
        legend = tk.Frame(left_frame, bg=self.theme['bg_secondary'])
        legend.pack(fill=tk.X, pady=2)
        tk.Label(legend, text="● Front", fg=self.theme['pad_front'],
                 bg=self.theme['bg_secondary'], font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=5)
        tk.Label(legend, text="● Back", fg=self.theme['pad_back'],
                 bg=self.theme['bg_secondary'], font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=5)
        tk.Label(legend, text="● Highlight", fg=self.theme['pad_highlight'],
                 bg=self.theme['bg_secondary'], font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=5)
        
        # ===== PANNEAU DROIT: LISTE =====
        right_frame = tk.Frame(self.paned, bg=self.theme['bg_primary'])
        self.paned.add(right_frame, width=650)
        
        # Toolbar Liste
        list_toolbar = tk.Frame(right_frame, bg=self.theme['bg_secondary'])
        list_toolbar.pack(fill=tk.X, pady=2)
        
        tk.Label(list_toolbar, text="Composants", font=('Segoe UI', 11, 'bold'),
                bg=self.theme['bg_secondary'], fg=self.theme['text_primary']).pack(side=tk.LEFT, padx=10)
        
        tk.Checkbutton(list_toolbar, text="Grouper", variable=self.group_by_value_var,
                       command=self._update_list, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=10)
        
        self.stats_label = tk.Label(list_toolbar, text="", font=('Segoe UI', 9),
                                   bg=self.theme['bg_secondary'], fg=self.theme['text_secondary'])
        self.stats_label.pack(side=tk.RIGHT, padx=10)
        
        # Treeview
        style = ttk.Style()
        style.configure("Split.Treeview", 
                       background=self.theme['bg_primary'],
                       foreground=self.theme['text_primary'],
                       fieldbackground=self.theme['bg_primary'],
                       font=('Segoe UI', 10))
        style.configure("Split.Treeview.Heading",
                       background=self.theme['bg_tertiary'],
                       foreground=self.theme['text_primary'],
                       font=('Segoe UI', 9, 'bold'))
        style.map('Split.Treeview', background=[('selected', self.theme['accent'])])
        
        columns = ('done', 'qty', 'ref', 'value', 'footprint', 'lcsc')
        self.tree = ttk.Treeview(right_frame, columns=columns, show='headings', 
                                 style="Split.Treeview")
        
        self.tree.tag_configure('done', background=self.theme['row_done'])
        self.tree.tag_configure('pending', background=self.theme['row_pending'])
        
        self.tree.heading('done', text='✓')
        self.tree.heading('qty', text='Qté')
        self.tree.heading('ref', text='Références')
        self.tree.heading('value', text='Valeur')
        self.tree.heading('footprint', text='Footprint')
        self.tree.heading('lcsc', text='LCSC')
        
        self.tree.column('done', width=30, anchor='center')
        self.tree.column('qty', width=35, anchor='center')
        self.tree.column('ref', width=150)
        self.tree.column('value', width=100)
        self.tree.column('footprint', width=130)
        self.tree.column('lcsc', width=80)
        
        scrollbar = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bindings liste
        self.tree.bind('<<TreeviewSelect>>', self._on_list_select)
        self.tree.bind('<Double-1>', self._on_toggle_processed)
        self.tree.bind('<space>', self._on_toggle_processed)
        
        # Barre d'actions en bas
        action_frame = tk.Frame(right_frame, bg=self.theme['bg_secondary'])
        action_frame.pack(fill=tk.X, pady=5)
        
        tk.Button(action_frame, text="✓ Marquer fait", command=self._mark_done, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="✗ Démarquer", command=self._mark_undone, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="◀ Préc", command=self._navigate_prev, **btn_style).pack(side=tk.LEFT, padx=5)
        tk.Button(action_frame, text="Suiv ▶", command=self._navigate_next, **btn_style).pack(side=tk.LEFT, padx=5)
        
        self.nav_label = tk.Label(action_frame, text="", bg=self.theme['bg_secondary'],
                                 fg=self.theme['text_secondary'], font=('Segoe UI', 9))
        self.nav_label.pack(side=tk.RIGHT, padx=10)
        
        # Initialiser la liste
        self._update_list()
    
    def _pcb_to_canvas(self, x, y):
        """Convertit coordonnées PCB -> canvas"""
        bbox = self.parser.board_bbox
        canvas_height = self.canvas.winfo_height() or 700
        cx = self.offset_x + (x - bbox['minx']) * self.scale
        cy = canvas_height - (self.offset_y + (y - bbox['miny']) * self.scale)
        return cx, cy
    
    def _draw_pcb(self, recalculate_scale=True):
        """Dessine le PCB avec highlight des composants sélectionnés"""
        self.canvas.delete('all')
        
        bbox = self.parser.board_bbox
        width = bbox['maxx'] - bbox['minx']
        height = bbox['maxy'] - bbox['miny']
        
        canvas_width = self.canvas.winfo_width() or 700
        canvas_height = self.canvas.winfo_height() or 700
        
        if recalculate_scale:
            scale_x = (canvas_width - 50) / width if width > 0 else 1
            scale_y = (canvas_height - 50) / height if height > 0 else 1
            self.scale = min(scale_x, scale_y) * 0.9
            self.offset_x = (canvas_width - width * self.scale) / 2
            self.offset_y = (canvas_height - height * self.scale) / 2
        
        # Fond
        x1, y1 = self._pcb_to_canvas(bbox['minx'], bbox['miny'])
        x2, y2 = self._pcb_to_canvas(bbox['maxx'], bbox['maxy'])
        self.canvas.create_rectangle(min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2),
                                     outline=self.theme['pcb_edge'], fill=self.theme['pcb_board'], width=2)
        
        # Tracks
        if self.show_tracks_var.get():
            for layer, layer_tracks in self.parser.tracks.items():
                if not isinstance(layer_tracks, list):
                    continue
                is_front = layer.startswith('F') or layer == 'F.Cu'
                color = self.theme['track_front'] if is_front else self.theme['track_back']
                for track in layer_tracks:
                    start = track.get('start')
                    end = track.get('end')
                    if start and end:
                        tx1, ty1 = self._pcb_to_canvas(start[0], start[1])
                        tx2, ty2 = self._pcb_to_canvas(end[0], end[1])
                        w = max(0.5, track.get('width', 0.2) * self.scale)
                        self.canvas.create_line(tx1, ty1, tx2, ty2, fill=color, width=w, capstyle=tk.ROUND)
        
        # Pads avec highlight
        if self.show_pads_var.get():
            for fp in self.parser.footprints:
                ref = fp.get('ref', '')
                fp_layer = fp.get('layer', 'F')
                is_highlighted = ref in self.highlighted_refs
                
                for pad in fp.get('pads', []):
                    self._draw_pad(pad, fp_layer, is_highlighted)
        
        # Silkscreen
        if self.show_silk_var.get():
            for fp in self.parser.footprints:
                ref = fp.get('ref', '')
                bbox_fp = fp.get('bbox', {})
                if ref and bbox_fp:
                    is_highlighted = ref in self.highlighted_refs
                    self._draw_ref(ref, bbox_fp, is_highlighted)
    
    def _draw_pad(self, pad, fp_layer, is_highlighted=False):
        """Dessine un pad"""
        pos = pad.get('pos', [0, 0])
        size = pad.get('size', [0.5, 0.5])
        shape = pad.get('shape', 'rect')
        layers = pad.get('layers', [fp_layer])
        
        cx, cy = self._pcb_to_canvas(pos[0], pos[1])
        w = max(2, size[0] * self.scale)
        h = max(2, size[1] * self.scale)
        
        if is_highlighted:
            color = self.theme['pad_highlight']
        else:
            is_front = 'F' in layers or any(l.startswith('F.') for l in layers)
            color = self.theme['pad_front'] if is_front else self.theme['pad_back']
        
        if shape == 'circle':
            r = w / 2
            self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r, fill=color, outline='')
        else:
            self.canvas.create_rectangle(cx - w/2, cy - h/2, cx + w/2, cy + h/2, fill=color, outline='')
    
    def _draw_ref(self, ref, bbox, is_highlighted=False):
        """Dessine la référence d'un composant"""
        if not ref or ref == 'REF**':
            return
        
        pos = bbox.get('pos', [0, 0])
        relpos = bbox.get('relpos', [0, 0])
        size = bbox.get('size', [1, 1])
        
        center_x = pos[0] + relpos[0] + size[0] / 2
        center_y = pos[1] + relpos[1] + size[1] / 2
        cx, cy = self._pcb_to_canvas(center_x, center_y)
        
        font_size = max(6, min(10, int(min(size[0], size[1]) * self.scale * 0.4)))
        color = self.theme['pad_highlight'] if is_highlighted else self.theme['silk_text']
        
        self.canvas.create_text(cx, cy, text=ref, fill=color, font=('Consolas', font_size, 'bold'))
    
    def _update_list(self):
        """Met à jour la liste des composants"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if self.group_by_value_var.get():
            grouped = {}
            for comp in self.components:
                key = (comp['value'], comp['footprint'], comp.get('lcsc', ''))
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(comp['ref'])
            
            for (value, footprint, lcsc), refs in sorted(grouped.items()):
                refs_sorted = sorted(refs, key=lambda r: (r[0], int(''.join(filter(str.isdigit, r)) or 0)))
                key = (value, footprint, lcsc)
                is_done = key in self.processed_items
                tag = 'done' if is_done else 'pending'
                
                self.tree.insert('', tk.END, values=(
                    '✓' if is_done else '', len(refs), ', '.join(refs_sorted),
                    value, footprint, lcsc
                ), tags=(tag,))
        else:
            for comp in sorted(self.components, key=lambda c: (c['value'], c['ref'])):
                key = (comp['value'], comp['footprint'], comp.get('lcsc', ''))
                is_done = key in self.processed_items
                tag = 'done' if is_done else 'pending'
                
                self.tree.insert('', tk.END, values=(
                    '✓' if is_done else '', 1, comp['ref'],
                    comp['value'], comp['footprint'], comp.get('lcsc', '')
                ), tags=(tag,))
        
        # Stats
        total = len(self.tree.get_children())
        done = sum(1 for item in self.tree.get_children() if self.tree.item(item, 'values')[0] == '✓')
        self.stats_label.config(text=f"{done}/{total} faits ({int(done/total*100) if total > 0 else 0}%)")
        self._update_nav_label()
    
    def _on_list_select(self, event=None):
        """Highlight les composants sélectionnés sur le PCB"""
        self.highlighted_refs.clear()
        
        for item in self.tree.selection():
            values = self.tree.item(item, 'values')
            if len(values) >= 3:
                refs_str = values[2]
                for ref in refs_str.split(', '):
                    ref = ref.strip()
                    if ref:
                        self.highlighted_refs.add(ref)
        
        self._draw_pcb(recalculate_scale=False)
    
    def _on_toggle_processed(self, event=None):
        """Bascule l'état traité"""
        for item in self.tree.selection():
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                if key in self.processed_items:
                    self.processed_items.discard(key)
                else:
                    self.processed_items.add(key)
        
        self._update_list()
        self.on_processed_change()
    
    def _mark_done(self):
        """Marque comme fait"""
        for item in self.tree.selection():
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                self.processed_items.add(key)
        self._update_list()
        self.on_processed_change()
    
    def _mark_undone(self):
        """Démarque"""
        for item in self.tree.selection():
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                self.processed_items.discard(key)
        self._update_list()
        self.on_processed_change()
    
    def _navigate_prev(self):
        """Navigue vers l'élément précédent"""
        children = self.tree.get_children()
        if not children:
            return
        
        selection = self.tree.selection()
        if selection:
            idx = children.index(selection[0])
            new_idx = (idx - 1) % len(children)
        else:
            new_idx = len(children) - 1
        
        self.tree.selection_set(children[new_idx])
        self.tree.see(children[new_idx])
        self._on_list_select()
        self._update_nav_label()
    
    def _navigate_next(self):
        """Navigue vers l'élément suivant"""
        children = self.tree.get_children()
        if not children:
            return
        
        selection = self.tree.selection()
        if selection:
            idx = children.index(selection[0])
            new_idx = (idx + 1) % len(children)
        else:
            new_idx = 0
        
        self.tree.selection_set(children[new_idx])
        self.tree.see(children[new_idx])
        self._on_list_select()
        self._update_nav_label()
    
    def _update_nav_label(self):
        """Met à jour le label de navigation"""
        children = self.tree.get_children()
        selection = self.tree.selection()
        if children and selection:
            idx = children.index(selection[0]) + 1
            self.nav_label.config(text=f"{idx}/{len(children)}")
        else:
            self.nav_label.config(text=f"0/{len(children)}")
    
    def _on_pan_start(self, event):
        self.pan_start_x = event.x
        self.pan_start_y = event.y
    
    def _on_pan_drag(self, event):
        if self.pan_start_x is None:
            return
        self.offset_x += event.x - self.pan_start_x
        self.offset_y -= event.y - self.pan_start_y
        self.pan_start_x = event.x
        self.pan_start_y = event.y
        self._draw_pcb(recalculate_scale=False)
    
    def _on_mousewheel(self, event):
        if event.delta > 0:
            self._zoom_in()
        else:
            self._zoom_out()
    
    def _zoom_in(self):
        self.scale *= 1.2
        self._draw_pcb(recalculate_scale=False)
    
    def _zoom_out(self):
        self.scale /= 1.2
        self._draw_pcb(recalculate_scale=False)
    
    def _reset_view(self):
        self._draw_pcb(recalculate_scale=True)


# ==================== MAIN APPLICATION ====================

class IBomSelectorApp:
    """Application principale avec interface moderne"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.prefs = Preferences()
        self.theme = THEMES[self.prefs.get('theme', 'dark')]
        
        self.root.title("IBom Component Selector v2.0")
        self.root.geometry("1200x900")
        self.root.configure(bg=self.theme['bg_primary'])
        
        self.parser = None
        self.selected_components = []
        self.filtered_components = []
        self.selection_rect = None
        self.processed_items = set()
        self.sort_column = None
        self.sort_reverse = False
        self.history = []
        self.history_file = None
        self.current_history_index = None
        self.current_item_index = 0  # Pour navigation
        self.view_mode = 'split'  # 'split', 'list', 'pcb'
        self.highlighted_refs = set()  # Pour highlight PCB
        self.pcb_scale = 1.0
        self.pcb_offset_x = 50
        self.pcb_offset_y = 50
        self.show_pads_var = None
        self.show_tracks_var = None
        self.show_silk_var = None
        
        # Variables
        self.layer_filter = tk.StringVar(value="all")
        self.search_var = tk.StringVar()
        self.group_by_value_var = tk.BooleanVar(value=self.prefs.get('group_by_value', True))
        self.status_filter = tk.StringVar(value="all")  # all, done, pending
        
        self._setup_ui()
        self._setup_keyboard_shortcuts()
        self._auto_load_bom()
    
    def _apply_theme(self):
        """Applique le thème à tous les widgets"""
        self.theme = THEMES[self.prefs.get('theme', 'dark')]
        self.root.configure(bg=self.theme['bg_primary'])
    
    def _toggle_theme(self):
        """Bascule entre thème sombre et clair"""
        current = self.prefs.get('theme', 'dark')
        new_theme = 'light' if current == 'dark' else 'dark'
        self.prefs.set('theme', new_theme)
        messagebox.showinfo("Thème", f"Thème changé en '{new_theme}'.\nRedémarrez l'application pour appliquer.")
    
    def _setup_ui(self):
        """Configure l'interface utilisateur principale"""
        # Frame principal
        main_frame = tk.Frame(self.root, bg=self.theme['bg_primary'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Header avec titre et boutons thème
        header_frame = tk.Frame(main_frame, bg=self.theme['bg_primary'])
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        title = tk.Label(
            header_frame,
            text="IBom Component Selector",
            font=('Segoe UI', 18, 'bold'),
            bg=self.theme['bg_primary'],
            fg=self.theme['text_primary']
        )
        title.pack(side=tk.LEFT)
        
        # Boutons du header
        btn_style = {'bg': self.theme['bg_tertiary'], 'fg': self.theme['text_primary'],
                     'activebackground': self.theme['accent'], 'activeforeground': '#ffffff',
                     'relief': tk.FLAT, 'padx': 10, 'pady': 3}
        
        tk.Button(header_frame, text="🌙 Thème", command=self._toggle_theme, **btn_style).pack(side=tk.RIGHT, padx=5)
        tk.Button(header_frame, text="⚙️ Options", command=self._show_options, **btn_style).pack(side=tk.RIGHT, padx=5)
        
        self.view_mode_btn = tk.Button(header_frame, text="📱 SPLIT", command=self._toggle_view_mode, 
                                       width=10, **btn_style)
        self.view_mode_btn.pack(side=tk.RIGHT, padx=5)
        
        # Frame pour le fichier
        file_frame = tk.LabelFrame(main_frame, text="Fichier HTML", 
                                   bg=self.theme['bg_secondary'], fg=self.theme['text_primary'],
                                   font=('Segoe UI', 10, 'bold'))
        file_frame.pack(fill=tk.X, pady=10)
        
        file_inner = tk.Frame(file_frame, bg=self.theme['bg_secondary'])
        file_inner.pack(fill=tk.X, padx=10, pady=10)
        
        self.file_var = tk.StringVar()
        file_entry = tk.Entry(file_inner, textvariable=self.file_var, width=70,
                             bg=self.theme['bg_primary'], fg=self.theme['text_primary'],
                             insertbackground=self.theme['text_primary'])
        file_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        tk.Button(file_inner, text="Parcourir...", command=self._browse_file, **btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(file_inner, text="Charger", command=self._load_file, **btn_style).pack(side=tk.LEFT, padx=2)
        
        # Barre d'actions rapides
        action_bar = tk.Frame(main_frame, bg=self.theme['bg_secondary'])
        action_bar.pack(fill=tk.X, pady=5)
        
        self.clear_btn = tk.Button(action_bar, text="Effacer sélection", command=self._clear_selection,
                                   state=tk.DISABLED, width=15, **btn_style)
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_btn = tk.Button(action_bar, text="Exporter Excel", command=self._export_excel,
                                    state=tk.DISABLED, width=12, **btn_style)
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        self.export_csv_btn = tk.Button(action_bar, text="Exporter CSV", command=self._export_csv,
                                        state=tk.DISABLED, width=12, **btn_style)
        self.export_csv_btn.pack(side=tk.LEFT, padx=5)
        
        self.status_var = tk.StringVar(value="Chargez un fichier HTML pour commencer")
        tk.Label(action_bar, textvariable=self.status_var, font=('Segoe UI', 9),
                bg=self.theme['bg_secondary'], fg=self.theme['text_secondary']).pack(side=tk.LEFT, padx=20)
        
        # Barre de progression
        tk.Label(action_bar, text="Progression:", bg=self.theme['bg_secondary'],
                fg=self.theme['text_primary'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(20, 5))
        
        self.progress_canvas = tk.Canvas(action_bar, width=200, height=18, bg=self.theme['progress_bg'],
                                         highlightthickness=0)
        self.progress_canvas.pack(side=tk.LEFT, padx=5)
        
        self.progress_label = tk.Label(action_bar, text="0%", bg=self.theme['bg_secondary'],
                                       fg=self.theme['text_primary'], font=('Segoe UI', 9, 'bold'))
        self.progress_label.pack(side=tk.LEFT)
        
        # ========== CONTENU PRINCIPAL: PanedWindow PCB/Liste ==========
        self.content_paned = tk.PanedWindow(main_frame, orient=tk.VERTICAL, bg=self.theme['bg_primary'],
                                           sashwidth=6, sashrelief=tk.RAISED)
        self.content_paned.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # ----- Panneau PCB -----
        self.pcb_frame = tk.LabelFrame(self.content_paned, text="PCB (clic = sélection zone, sélection liste = highlight)",
                                       bg=self.theme['bg_secondary'], fg=self.theme['text_primary'],
                                       font=('Segoe UI', 10, 'bold'))
        self.content_paned.add(self.pcb_frame, height=350)
        
        # Toolbar PCB
        pcb_toolbar = tk.Frame(self.pcb_frame, bg=self.theme['bg_secondary'])
        pcb_toolbar.pack(fill=tk.X, padx=5, pady=2)
        
        self.show_pads_var = tk.BooleanVar(value=True)
        self.show_tracks_var = tk.BooleanVar(value=True)
        self.show_silk_var = tk.BooleanVar(value=True)
        
        tk.Checkbutton(pcb_toolbar, text="Pads", variable=self.show_pads_var,
                       command=self._draw_main_pcb, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(pcb_toolbar, text="Tracks", variable=self.show_tracks_var,
                       command=self._draw_main_pcb, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(pcb_toolbar, text="Silk", variable=self.show_silk_var,
                       command=self._draw_main_pcb, bg=self.theme['bg_secondary'],
                       fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary']
                       ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(pcb_toolbar, text="Reset", command=self._reset_pcb_view, width=6, **btn_style).pack(side=tk.RIGHT, padx=2)
        tk.Button(pcb_toolbar, text="Zoom -", command=self._zoom_out_pcb, width=6, **btn_style).pack(side=tk.RIGHT, padx=2)
        tk.Button(pcb_toolbar, text="Zoom +", command=self._zoom_in_pcb, width=6, **btn_style).pack(side=tk.RIGHT, padx=2)
        
        # Légende
        tk.Label(pcb_toolbar, text="●", fg=self.theme['pad_front'], bg=self.theme['bg_secondary'],
                font=('Segoe UI', 10)).pack(side=tk.RIGHT, padx=(10, 0))
        tk.Label(pcb_toolbar, text="Front", fg=self.theme['text_secondary'], bg=self.theme['bg_secondary'],
                font=('Segoe UI', 8)).pack(side=tk.RIGHT)
        tk.Label(pcb_toolbar, text="●", fg=self.theme['pad_back'], bg=self.theme['bg_secondary'],
                font=('Segoe UI', 10)).pack(side=tk.RIGHT, padx=(10, 0))
        tk.Label(pcb_toolbar, text="Back", fg=self.theme['text_secondary'], bg=self.theme['bg_secondary'],
                font=('Segoe UI', 8)).pack(side=tk.RIGHT)
        tk.Label(pcb_toolbar, text="●", fg=self.theme['pad_highlight'], bg=self.theme['bg_secondary'],
                font=('Segoe UI', 10)).pack(side=tk.RIGHT, padx=(10, 0))
        tk.Label(pcb_toolbar, text="Sélection", fg=self.theme['text_secondary'], bg=self.theme['bg_secondary'],
                font=('Segoe UI', 8)).pack(side=tk.RIGHT)
        
        # Canvas PCB principal
        self.pcb_canvas = tk.Canvas(self.pcb_frame, bg=self.theme['pcb_board'], highlightthickness=0)
        self.pcb_canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.pcb_canvas.bind('<Button-1>', self._on_pcb_click)
        self.pcb_canvas.bind('<MouseWheel>', self._on_pcb_mousewheel)
        self.pcb_canvas.bind('<Button-3>', self._on_pcb_pan_start)
        self.pcb_canvas.bind('<B3-Motion>', self._on_pcb_pan_drag)
        self.pcb_pan_start_x = None
        self.pcb_pan_start_y = None
        
        # ----- Panneau Liste -----
        self.list_frame = tk.LabelFrame(self.content_paned, text="Composants (double-clic = marquer fait)",
                                        bg=self.theme['bg_secondary'], fg=self.theme['text_primary'],
                                        font=('Segoe UI', 10, 'bold'))
        self.content_paned.add(self.list_frame, height=350)
        
        # Toolbar Liste (filtres)
        list_toolbar = tk.Frame(self.list_frame, bg=self.theme['bg_secondary'])
        list_toolbar.pack(fill=tk.X, padx=5, pady=2)
        
        tk.Label(list_toolbar, text="Couche:", bg=self.theme['bg_secondary'],
                fg=self.theme['text_primary'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        for text, value in [("All", "all"), ("F", "F"), ("B", "B")]:
            tk.Radiobutton(list_toolbar, text=text, variable=self.layer_filter, value=value,
                          command=self._apply_filters, bg=self.theme['bg_secondary'],
                          fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary'],
                          font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(list_toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        tk.Label(list_toolbar, text="Statut:", bg=self.theme['bg_secondary'],
                fg=self.theme['text_primary'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        for text, value in [("Tous", "all"), ("✓", "done"), ("○", "pending")]:
            tk.Radiobutton(list_toolbar, text=text, variable=self.status_filter, value=value,
                          command=self._apply_filters, bg=self.theme['bg_secondary'],
                          fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary'],
                          font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(list_toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        self.search_entry = tk.Entry(list_toolbar, textvariable=self.search_var, width=15,
                                    bg=self.theme['bg_primary'], fg=self.theme['text_primary'],
                                    insertbackground=self.theme['text_primary'], font=('Segoe UI', 9))
        self.search_entry.pack(side=tk.LEFT, padx=2)
        self.search_var.trace('w', lambda *args: self._apply_filters())
        
        tk.Button(list_toolbar, text="✕", width=2, command=lambda: self.search_var.set(""),
                 bg=self.theme['bg_tertiary'], fg=self.theme['text_primary'], relief=tk.FLAT
                 ).pack(side=tk.LEFT, padx=2)
        
        tk.Checkbutton(list_toolbar, text="Grouper", variable=self.group_by_value_var,
                      command=self._apply_filters, bg=self.theme['bg_secondary'],
                      fg=self.theme['text_primary'], selectcolor=self.theme['bg_tertiary'],
                      font=('Segoe UI', 8)).pack(side=tk.LEFT, padx=5)
        
        self.stats_var = tk.StringVar()
        tk.Label(list_toolbar, textvariable=self.stats_var, bg=self.theme['bg_secondary'],
                fg=self.theme['text_secondary'], font=('Segoe UI', 8, 'italic')).pack(side=tk.RIGHT, padx=5)
        
        # Historique
        self.history_var = tk.StringVar()
        self.history_combo = ttk.Combobox(list_toolbar, textvariable=self.history_var, 
                                           state='readonly', width=20, font=('Segoe UI', 8))
        self.history_combo.pack(side=tk.RIGHT, padx=2)
        tk.Label(list_toolbar, text="Hist:", bg=self.theme['bg_secondary'],
                fg=self.theme['text_primary'], font=('Segoe UI', 8)).pack(side=tk.RIGHT)
        
        # Container pour tree + scrollbar
        tree_container = tk.Frame(self.list_frame, bg=self.theme['bg_primary'])
        tree_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Treeview
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview", 
                       background=self.theme['bg_primary'],
                       foreground=self.theme['text_primary'],
                       fieldbackground=self.theme['bg_primary'],
                       font=('Segoe UI', self.prefs.get('font_size', 11)))
        style.configure("Treeview.Heading",
                       background=self.theme['bg_tertiary'],
                       foreground=self.theme['text_primary'],
                       font=('Segoe UI', 10, 'bold'))
        style.map('Treeview', background=[('selected', self.theme['accent'])])
        
        columns = ('done', 'qty', 'ref', 'value', 'footprint', 'lcsc')
        self.tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=10)
        
        self.tree.tag_configure('done', background=self.theme['row_done'])
        self.tree.tag_configure('pending', background=self.theme['row_pending'])
        self.tree.tag_configure('current', background=self.theme['accent'])
        
        self.tree.heading('done', text='✓', command=lambda: self._sort_by_column('done'))
        self.tree.heading('qty', text='Qté ↕', command=lambda: self._sort_by_column('qty'))
        self.tree.heading('ref', text='Références ↕', command=lambda: self._sort_by_column('ref'))
        self.tree.heading('value', text='Valeur ↕', command=lambda: self._sort_by_column('value'))
        self.tree.heading('footprint', text='Footprint ↕', command=lambda: self._sort_by_column('footprint'))
        self.tree.heading('lcsc', text='LCSC ↕', command=lambda: self._sort_by_column('lcsc'))
        
        self.tree.column('done', width=40, anchor='center')
        self.tree.column('qty', width=50, anchor='center')
        self.tree.column('ref', width=180)
        self.tree.column('value', width=120)
        self.tree.column('footprint', width=180)
        self.tree.column('lcsc', width=100)
        
        self.tree.bind('<Double-1>', self._toggle_processed)
        self.tree.bind('<space>', self._toggle_processed)
        self.tree.bind('<<TreeviewSelect>>', self._on_tree_select)
        
        scrollbar = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Boutons de navigation et actions en bas de la liste
        nav_frame = tk.Frame(self.list_frame, bg=self.theme['bg_secondary'])
        nav_frame.pack(fill=tk.X, padx=5, pady=2)
        
        # Navigation
        nav_btn_style = {'bg': self.theme['bg_tertiary'], 'fg': self.theme['text_primary'],
                        'activebackground': self.theme['accent'], 'activeforeground': '#ffffff',
                        'relief': tk.FLAT, 'padx': 10, 'pady': 2, 'font': ('Segoe UI', 9)}
        
        tk.Button(nav_frame, text="◀ Préc", command=self._navigate_prev, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(nav_frame, text="Suiv ▶", command=self._navigate_next, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(nav_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        
        tk.Button(nav_frame, text="✓ Fait", command=self._mark_selected_processed, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(nav_frame, text="✗ Défaire", command=self._unmark_selected_processed, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(nav_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)
        
        tk.Button(nav_frame, text="Sauver", command=self._save_current_to_history, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        tk.Button(nav_frame, text="Charger", command=self._load_history_selection, **nav_btn_style).pack(side=tk.LEFT, padx=2)
        
        self.nav_label = tk.Label(nav_frame, text="", bg=self.theme['bg_secondary'],
                                 fg=self.theme['text_secondary'], font=('Segoe UI', 9))
        self.nav_label.pack(side=tk.RIGHT, padx=5)
    
    def _toggle_view_mode(self):
        """Bascule entre les modes SPLIT, LIST, PCB"""
        modes = ['split', 'list', 'pcb']
        current_idx = modes.index(self.view_mode)
        self.view_mode = modes[(current_idx + 1) % 3]
        
        # Mettre à jour le bouton
        mode_icons = {'split': '📱 SPLIT', 'list': '📋 LIST', 'pcb': '🔌 PCB'}
        self.view_mode_btn.config(text=mode_icons[self.view_mode])
        
        # Mettre à jour la visibilité
        self._update_view_layout()
    
    def _update_view_layout(self):
        """Met à jour le layout selon le mode"""
        if self.view_mode == 'split':
            # Les deux panneaux visibles
            self.content_paned.paneconfig(self.pcb_frame, height=350)
            self.content_paned.paneconfig(self.list_frame, height=350)
            try:
                self.content_paned.add(self.pcb_frame)
            except:
                pass
            try:
                self.content_paned.add(self.list_frame)
            except:
                pass
        elif self.view_mode == 'list':
            # Seulement la liste
            try:
                self.content_paned.forget(self.pcb_frame)
            except:
                pass
            try:
                self.content_paned.add(self.list_frame)
            except:
                pass
        elif self.view_mode == 'pcb':
            # Seulement le PCB
            try:
                self.content_paned.forget(self.list_frame)
            except:
                pass
            try:
                self.content_paned.add(self.pcb_frame)
            except:
                pass
        
        # Redessiner le PCB après changement
        self.root.after(100, self._draw_main_pcb)
    
    def _on_tree_select(self, event=None):
        """Highlight les composants sélectionnés sur le PCB"""
        self.highlighted_refs.clear()
        
        for item in self.tree.selection():
            values = self.tree.item(item, 'values')
            if len(values) >= 3:
                refs_str = values[2]
                for ref in refs_str.split(', '):
                    ref = ref.strip()
                    if ref:
                        self.highlighted_refs.add(ref)
        
        self._draw_main_pcb(recalculate_scale=False)
        self._update_nav_label()
    
    # ========== MÉTHODES PCB PRINCIPAL ==========
    
    def _pcb_to_canvas_main(self, x, y):
        """Convertit coordonnées PCB -> canvas principal"""
        if not self.parser:
            return 0, 0
        bbox = self.parser.board_bbox
        canvas_height = self.pcb_canvas.winfo_height() or 300
        cx = self.pcb_offset_x + (x - bbox['minx']) * self.pcb_scale
        cy = canvas_height - (self.pcb_offset_y + (y - bbox['miny']) * self.pcb_scale)
        return cx, cy
    
    def _canvas_to_pcb_main(self, canvas_x, canvas_y):
        """Convertit canvas -> PCB"""
        if not self.parser:
            return 0, 0
        bbox = self.parser.board_bbox
        canvas_height = self.pcb_canvas.winfo_height() or 300
        x = (canvas_x - self.pcb_offset_x) / self.pcb_scale + bbox['minx']
        y = (canvas_height - canvas_y - self.pcb_offset_y) / self.pcb_scale + bbox['miny']
        return x, y
    
    def _draw_main_pcb(self, recalculate_scale=True):
        """Dessine le PCB principal avec highlight"""
        self.pcb_canvas.delete('all')
        
        if not self.parser:
            self.pcb_canvas.create_text(
                self.pcb_canvas.winfo_width() // 2 or 300,
                self.pcb_canvas.winfo_height() // 2 or 150,
                text="Chargez un fichier pour voir le PCB",
                fill=self.theme['text_secondary'], font=('Segoe UI', 12)
            )
            return
        
        bbox = self.parser.board_bbox
        width = bbox['maxx'] - bbox['minx']
        height = bbox['maxy'] - bbox['miny']
        
        canvas_width = self.pcb_canvas.winfo_width() or 600
        canvas_height = self.pcb_canvas.winfo_height() or 300
        
        if recalculate_scale:
            scale_x = (canvas_width - 40) / width if width > 0 else 1
            scale_y = (canvas_height - 40) / height if height > 0 else 1
            self.pcb_scale = min(scale_x, scale_y) * 0.9
            self.pcb_offset_x = (canvas_width - width * self.pcb_scale) / 2
            self.pcb_offset_y = (canvas_height - height * self.pcb_scale) / 2
        
        # Fond
        x1, y1 = self._pcb_to_canvas_main(bbox['minx'], bbox['miny'])
        x2, y2 = self._pcb_to_canvas_main(bbox['maxx'], bbox['maxy'])
        self.pcb_canvas.create_rectangle(min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2),
                                         outline=self.theme['pcb_edge'], fill=self.theme['pcb_board'], width=2)
        
        # Tracks
        if self.show_tracks_var and self.show_tracks_var.get():
            for layer, layer_tracks in self.parser.tracks.items():
                if not isinstance(layer_tracks, list):
                    continue
                is_front = layer.startswith('F') or layer == 'F.Cu'
                color = self.theme['track_front'] if is_front else self.theme['track_back']
                for track in layer_tracks:
                    start = track.get('start')
                    end = track.get('end')
                    if start and end:
                        tx1, ty1 = self._pcb_to_canvas_main(start[0], start[1])
                        tx2, ty2 = self._pcb_to_canvas_main(end[0], end[1])
                        w = max(0.5, track.get('width', 0.2) * self.pcb_scale)
                        self.pcb_canvas.create_line(tx1, ty1, tx2, ty2, fill=color, width=w, capstyle=tk.ROUND)
        
        # Pads avec highlight
        if self.show_pads_var and self.show_pads_var.get():
            for fp in self.parser.footprints:
                ref = fp.get('ref', '')
                fp_layer = fp.get('layer', 'F')
                is_highlighted = ref in self.highlighted_refs
                
                for pad in fp.get('pads', []):
                    self._draw_main_pad(pad, fp_layer, is_highlighted)
        
        # Silkscreen + refs
        if self.show_silk_var and self.show_silk_var.get():
            for fp in self.parser.footprints:
                ref = fp.get('ref', '')
                bbox_fp = fp.get('bbox', {})
                if ref and bbox_fp:
                    is_highlighted = ref in self.highlighted_refs
                    self._draw_main_ref(ref, bbox_fp, is_highlighted)
        
        # Zone de sélection
        if self.selection_rect:
            sx1, sy1, sx2, sy2 = self.selection_rect
            cx1, cy1 = self._pcb_to_canvas_main(sx1, sy1)
            cx2, cy2 = self._pcb_to_canvas_main(sx2, sy2)
            self.pcb_canvas.create_rectangle(min(cx1, cx2), min(cy1, cy2), max(cx1, cx2), max(cy1, cy2),
                                            outline=self.theme['selection_rect'], width=2, dash=(5, 3))
    
    def _draw_main_pad(self, pad, fp_layer, is_highlighted=False):
        """Dessine un pad sur le canvas principal"""
        pos = pad.get('pos', [0, 0])
        size = pad.get('size', [0.5, 0.5])
        shape = pad.get('shape', 'rect')
        layers = pad.get('layers', [fp_layer])
        
        cx, cy = self._pcb_to_canvas_main(pos[0], pos[1])
        w = max(2, size[0] * self.pcb_scale)
        h = max(2, size[1] * self.pcb_scale)
        
        if is_highlighted:
            color = self.theme['pad_highlight']
        else:
            is_front = 'F' in layers or any(l.startswith('F.') for l in layers)
            color = self.theme['pad_front'] if is_front else self.theme['pad_back']
        
        if shape == 'circle':
            r = w / 2
            self.pcb_canvas.create_oval(cx - r, cy - r, cx + r, cy + r, fill=color, outline='')
        else:
            self.pcb_canvas.create_rectangle(cx - w/2, cy - h/2, cx + w/2, cy + h/2, fill=color, outline='')
    
    def _draw_main_ref(self, ref, bbox, is_highlighted=False):
        """Dessine la référence d'un composant"""
        if not ref or ref == 'REF**':
            return
        
        pos = bbox.get('pos', [0, 0])
        relpos = bbox.get('relpos', [0, 0])
        size = bbox.get('size', [1, 1])
        
        center_x = pos[0] + relpos[0] + size[0] / 2
        center_y = pos[1] + relpos[1] + size[1] / 2
        cx, cy = self._pcb_to_canvas_main(center_x, center_y)
        
        font_size = max(5, min(9, int(min(size[0], size[1]) * self.pcb_scale * 0.4)))
        color = self.theme['pad_highlight'] if is_highlighted else self.theme['silk_text']
        
        self.pcb_canvas.create_text(cx, cy, text=ref, fill=color, font=('Consolas', font_size, 'bold'))
    
    def _zoom_in_pcb(self):
        self.pcb_scale *= 1.2
        self._draw_main_pcb(recalculate_scale=False)
    
    def _zoom_out_pcb(self):
        self.pcb_scale /= 1.2
        self._draw_main_pcb(recalculate_scale=False)
    
    def _reset_pcb_view(self):
        self._draw_main_pcb(recalculate_scale=True)
    
    def _on_pcb_mousewheel(self, event):
        if event.delta > 0:
            self._zoom_in_pcb()
        else:
            self._zoom_out_pcb()
    
    def _on_pcb_pan_start(self, event):
        self.pcb_pan_start_x = event.x
        self.pcb_pan_start_y = event.y
    
    def _on_pcb_pan_drag(self, event):
        if self.pcb_pan_start_x is None:
            return
        self.pcb_offset_x += event.x - self.pcb_pan_start_x
        self.pcb_offset_y -= event.y - self.pcb_pan_start_y
        self.pcb_pan_start_x = event.x
        self.pcb_pan_start_y = event.y
        self._draw_main_pcb(recalculate_scale=False)
    
    def _show_split_view(self):
        """Affiche la vue split PCB/Liste dans une fenêtre séparée"""
        if not self.selected_components:
            messagebox.showwarning("Attention", "Sélectionnez d'abord des composants sur le PCB")
            return
        
        if self.split_window and self.split_window.winfo_exists():
            self.split_window.lift()
            self.split_window.focus_force()
            return
        
        def on_processed_change():
            """Callback quand les items traités changent"""
            self._update_tree()
            self._update_progress()
        
        self.split_window = SplitView(
            self.root, 
            self.parser, 
            self.selected_components,
            self.processed_items,
            self.prefs,
            self.theme,
            on_processed_change
        )
        self.split_window.transient(self.root)
    
    def _show_options(self):
        """Affiche la fenêtre d'options"""
        options_win = tk.Toplevel(self.root)
        options_win.title("Options")
        options_win.geometry("400x300")
        options_win.configure(bg=self.theme['bg_primary'])
        options_win.transient(self.root)
        options_win.grab_set()
        
        # Taille de police
        font_frame = tk.Frame(options_win, bg=self.theme['bg_primary'])
        font_frame.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(font_frame, text="Taille police:", bg=self.theme['bg_primary'],
                fg=self.theme['text_primary']).pack(side=tk.LEFT)
        
        font_size_var = tk.IntVar(value=self.prefs.get('font_size', 11))
        font_sizes = [9, 10, 11, 12, 13, 14, 15]
        font_combo = ttk.Combobox(font_frame, textvariable=font_size_var, values=font_sizes, width=5)
        font_combo.pack(side=tk.LEFT, padx=10)
        
        # Sauvegarde auto
        auto_save_frame = tk.Frame(options_win, bg=self.theme['bg_primary'])
        auto_save_frame.pack(fill=tk.X, padx=20, pady=10)
        
        auto_save_var = tk.BooleanVar(value=self.prefs.get('auto_save', False))
        tk.Checkbutton(auto_save_frame, text="Sauvegarde automatique", variable=auto_save_var,
                      bg=self.theme['bg_primary'], fg=self.theme['text_primary'],
                      selectcolor=self.theme['bg_secondary']).pack(side=tk.LEFT)
        
        # Bouton sauvegarder
        def save_options():
            self.prefs.set('font_size', font_size_var.get())
            self.prefs.set('auto_save', auto_save_var.get())
            messagebox.showinfo("Options", "Options sauvegardées.")
            options_win.destroy()
        
        tk.Button(options_win, text="Sauvegarder", command=save_options,
                 bg=self.theme['success'], fg='#ffffff', relief=tk.FLAT,
                 padx=20, pady=5).pack(pady=20)
    
    def _update_progress(self):
        """Met à jour la barre de progression"""
        self.progress_canvas.delete('all')
        
        if not self.tree.get_children():
            self.progress_label.config(text="0%")
            return
        
        total = len(self.tree.get_children())
        done = sum(1 for item in self.tree.get_children() 
                  if self.tree.item(item, 'values')[0] == '✓')
        
        width = self.progress_canvas.winfo_width()
        height = self.progress_canvas.winfo_height()
        
        if total > 0:
            progress = done / total
            fill_width = width * progress
            
            self.progress_canvas.create_rectangle(0, 0, fill_width, height,
                                                  fill=self.theme['progress_fill'], outline='')
            
            percent = int(progress * 100)
            self.progress_label.config(text=f"{percent}% ({done}/{total})")
        else:
            self.progress_label.config(text="0%")
    
    def _navigate_next(self):
        """Navigue vers le composant suivant"""
        children = self.tree.get_children()
        if not children:
            return
        
        self.current_item_index = (self.current_item_index + 1) % len(children)
        item = children[self.current_item_index]
        
        self.tree.selection_set(item)
        self.tree.see(item)
        self._update_nav_label()
    
    def _navigate_prev(self):
        """Navigue vers le composant précédent"""
        children = self.tree.get_children()
        if not children:
            return
        
        self.current_item_index = (self.current_item_index - 1) % len(children)
        item = children[self.current_item_index]
        
        self.tree.selection_set(item)
        self.tree.see(item)
        self._update_nav_label()
    
    def _update_nav_label(self):
        """Met à jour le label de navigation"""
        children = self.tree.get_children()
        if children:
            self.nav_label.config(text=f"Composant {self.current_item_index + 1} / {len(children)}")
        else:
            self.nav_label.config(text="")
    
    def _browse_file(self):
        """Ouvre le dialogue de sélection de fichier"""
        filename = filedialog.askopenfilename(
            title="Sélectionner un fichier InteractiveHtmlBom",
            filetypes=[("Fichiers HTML", "*.html"), ("Tous les fichiers", "*.*")]
        )
        if filename:
            self.file_var.set(filename)
    
    def _load_file(self):
        """Charge et parse le fichier HTML"""
        filepath = self.file_var.get()
        if not filepath:
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier HTML")
            return
        
        if not Path(filepath).exists():
            messagebox.showerror("Erreur", "Le fichier n'existe pas")
            return
        
        try:
            self.parser = IBomParser(filepath)
            self.parser.parse()
            
            self.status_var.set(f"Chargé: {len(self.parser.components)} composants")
            self._load_history()
            self._draw_main_pcb()
            
            messagebox.showinfo(
                "Succès",
                f"Fichier chargé avec succès!\n"
                f"Composants: {len(self.parser.components)}\n"
                f"Footprints: {len(self.parser.footprints)}\n"
                f"Historique: {len(self.history)} sélections"
            )
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement:\n{str(e)}")
            self.status_var.set("Erreur lors du chargement")
    
    def _on_pcb_click(self, event):
        """Ouvre le viewer PCB"""
        if self.parser:
            viewer = PCBViewer(self.root, self.parser, self._on_selection, self.prefs, self.theme)
            viewer.transient(self.root)
            viewer.grab_set()
    
    def _on_selection(self, selected_components, selection_rect=None):
        """Callback de sélection"""
        self.selected_components = selected_components
        self.selection_rect = selection_rect
        self.current_item_index = 0
        
        self._apply_filters()
        self._draw_main_pcb()
        
        self.export_btn.config(state=tk.NORMAL)
        self.export_csv_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)
        self.status_var.set(f"{len(selected_components)} composants sélectionnés")
    
    def _apply_filters(self):
        """Applique les filtres"""
        layer_filter = self.layer_filter.get()
        status_filter = self.status_filter.get()
        search_text = self.search_var.get().lower().strip()
        
        self.filtered_components = []
        for comp in self.selected_components:
            # Filtre couche
            if layer_filter != "all" and comp.get('layer', 'F') != layer_filter:
                continue
            
            # Filtre recherche
            if search_text:
                searchable = f"{comp['ref']} {comp['value']} {comp['footprint']} {comp['lcsc']}".lower()
                if search_text not in searchable:
                    continue
            
            self.filtered_components.append(comp)
        
        self._update_tree()
        self._update_statistics()
        self._update_progress()
        self._update_nav_label()
    
    def _update_statistics(self):
        """Met à jour les statistiques"""
        if not self.selected_components:
            self.stats_var.set("")
            return
        
        total = len(self.selected_components)
        filtered = len(self.filtered_components)
        front = sum(1 for c in self.selected_components if c.get('layer', 'F') == 'F')
        back = total - front
        
        if filtered == total:
            self.stats_var.set(f"Total: {total} | Front: {front} | Back: {back}")
        else:
            self.stats_var.set(f"Affichés: {filtered}/{total} | Front: {front} | Back: {back}")
    
    def _update_tree(self):
        """Met à jour l'affichage de la liste"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        status_filter = self.status_filter.get()
        
        if self.group_by_value_var.get():
            # Regrouper
            grouped = {}
            for comp in self.filtered_components:
                key = (comp['value'], comp['footprint'], comp['lcsc'])
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(comp['ref'])
            
            data_list = []
            for (value, footprint, lcsc), refs in grouped.items():
                refs_sorted = sorted(refs, key=lambda r: (r[0], int(''.join(filter(str.isdigit, r)) or 0)))
                refs_str = ', '.join(refs_sorted)
                key = (value, footprint, lcsc)
                is_done = key in self.processed_items
                
                # Filtre statut
                if status_filter == 'done' and not is_done:
                    continue
                if status_filter == 'pending' and is_done:
                    continue
                
                data_list.append({
                    'key': key,
                    'done': '✓' if is_done else '',
                    'qty': len(refs),
                    'ref': refs_str,
                    'value': value,
                    'footprint': footprint,
                    'lcsc': lcsc,
                    'is_done': is_done
                })
        else:
            # Sans groupement
            data_list = []
            for comp in self.filtered_components:
                key = (comp['value'], comp['footprint'], comp['lcsc'])
                is_done = key in self.processed_items
                
                if status_filter == 'done' and not is_done:
                    continue
                if status_filter == 'pending' and is_done:
                    continue
                
                data_list.append({
                    'key': key,
                    'done': '✓' if is_done else '',
                    'qty': 1,
                    'ref': comp['ref'],
                    'value': comp['value'],
                    'footprint': comp['footprint'],
                    'lcsc': comp['lcsc'],
                    'is_done': is_done
                })
        
        # Tri
        if self.sort_column:
            if self.sort_column == 'done':
                data_list.sort(key=lambda x: (not x['is_done'], x['value']), reverse=self.sort_reverse)
            elif self.sort_column == 'qty':
                data_list.sort(key=lambda x: x['qty'], reverse=self.sort_reverse)
            elif self.sort_column == 'ref':
                data_list.sort(key=lambda x: x['ref'], reverse=self.sort_reverse)
            elif self.sort_column == 'value':
                data_list.sort(key=lambda x: x['value'], reverse=self.sort_reverse)
            elif self.sort_column == 'footprint':
                data_list.sort(key=lambda x: x['footprint'], reverse=self.sort_reverse)
            elif self.sort_column == 'lcsc':
                data_list.sort(key=lambda x: x['lcsc'], reverse=self.sort_reverse)
        else:
            data_list.sort(key=lambda x: (x['value'], x['ref']))
        
        for data in data_list:
            tag = 'done' if data['is_done'] else 'pending'
            self.tree.insert('', tk.END, values=(
                data['done'], data['qty'], data['ref'],
                data['value'], data['footprint'], data['lcsc']
            ), tags=(tag,))
    
    def _sort_by_column(self, column):
        """Trie par colonne"""
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        
        for col in ('done', 'qty', 'ref', 'value', 'footprint', 'lcsc'):
            text = self.tree.heading(col)['text'].rstrip(' ↑↓↕')
            if col == column:
                arrow = ' ↓' if self.sort_reverse else ' ↑'
            else:
                arrow = ' ↕' if col != 'done' else ''
            self.tree.heading(col, text=text + arrow)
        
        self._update_tree()
    
    def _toggle_processed(self, event=None):
        """Bascule l'état traité"""
        selection = self.tree.selection()
        if not selection:
            return
        
        for item in selection:
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                if key in self.processed_items:
                    self.processed_items.discard(key)
                else:
                    self.processed_items.add(key)
        
        self._update_tree()
        self._update_progress()
    
    def _mark_selected_processed(self):
        """Marque comme traité"""
        selection = self.tree.selection()
        for item in selection:
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                self.processed_items.add(key)
        self._update_tree()
        self._update_progress()
    
    def _unmark_selected_processed(self):
        """Démarque"""
        selection = self.tree.selection()
        for item in selection:
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                self.processed_items.discard(key)
        self._update_tree()
        self._update_progress()
    
    def _unmark_all_processed(self):
        """Tout démarquer"""
        self.processed_items.clear()
        self._update_tree()
        self._update_progress()
    
    def _clear_selection(self):
        """Efface la sélection"""
        self.selected_components = []
        self.filtered_components = []
        self.selection_rect = None
        self.processed_items.clear()
        self.current_item_index = 0
        self._update_tree()
        self._update_statistics()
        self._update_progress()
        self._update_nav_label()
        self._draw_main_pcb()
        self.export_btn.config(state=tk.DISABLED)
        self.export_csv_btn.config(state=tk.DISABLED)
        self.clear_btn.config(state=tk.DISABLED)
        self.status_var.set("Sélection effacée")
    
    def _export_excel(self):
        """Export Excel"""
        if not self.filtered_components:
            messagebox.showwarning("Attention", "Aucun composant à exporter")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")]
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM Sélection"
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ['Quantité', 'Référence', 'Valeur', 'Footprint', 'LCSC']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            grouped = {}
            for comp in self.filtered_components:
                key = (comp['value'], comp['footprint'], comp['lcsc'])
                if key not in grouped:
                    grouped[key] = {'refs': [], 'value': comp['value'],
                                   'footprint': comp['footprint'], 'lcsc': comp['lcsc']}
                grouped[key]['refs'].append(comp['ref'])
            
            row = 2
            for key, data in sorted(grouped.items()):
                refs = sorted(data['refs'])
                ws.cell(row=row, column=1, value=len(refs)).border = thin_border
                ws.cell(row=row, column=2, value=', '.join(refs)).border = thin_border
                ws.cell(row=row, column=3, value=data['value']).border = thin_border
                ws.cell(row=row, column=4, value=data['footprint']).border = thin_border
                ws.cell(row=row, column=5, value=data['lcsc']).border = thin_border
                row += 1
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            
            wb.save(filename)
            messagebox.showinfo("Succès", f"Fichier Excel créé!\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export:\n{str(e)}")
    
    def _export_csv(self):
        """Export CSV"""
        if not self.filtered_components:
            messagebox.showwarning("Attention", "Aucun composant à exporter")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Enregistrer le fichier CSV",
            defaultextension=".csv",
            filetypes=[("Fichiers CSV", "*.csv")]
        )
        
        if not filename:
            return
        
        try:
            grouped = {}
            for comp in self.filtered_components:
                key = (comp['value'], comp['footprint'], comp['lcsc'])
                if key not in grouped:
                    grouped[key] = {'refs': [], 'value': comp['value'],
                                   'footprint': comp['footprint'], 'lcsc': comp['lcsc']}
                grouped[key]['refs'].append(comp['ref'])
            
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Quantité', 'Référence', 'Valeur', 'Footprint', 'LCSC'])
                
                for key, data in sorted(grouped.items()):
                    refs = sorted(data['refs'])
                    writer.writerow([len(refs), ', '.join(refs), data['value'],
                                   data['footprint'], data['lcsc']])
            
            messagebox.showinfo("Succès", f"Fichier CSV créé!\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export:\n{str(e)}")
    
    # ==================== HISTORIQUE ====================
    
    def _get_history_file_path(self):
        if not self.file_var.get():
            return None
        html_path = Path(self.file_var.get())
        return html_path.parent / f".{html_path.stem}_history.json"
    
    def _load_history(self):
        self.history = []
        self.current_history_index = None
        self.history_file = self._get_history_file_path()
        
        if self.history_file and self.history_file.exists():
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    self.history = json.load(f)
            except Exception as e:
                print(f"Erreur chargement historique: {e}")
        
        self._update_history_combo()
    
    def _save_history(self):
        if not self.history_file:
            return
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erreur sauvegarde historique: {e}")
    
    def _update_history_combo(self):
        items = []
        for i, entry in enumerate(self.history):
            name = entry.get('name', f"Sélection {i+1}")
            date = entry.get('date', '')
            count = len(entry.get('components', []))
            processed = len(entry.get('processed', []))
            items.append(f"{name} ({count} comp., {processed} faits) - {date}")
        
        self.history_combo['values'] = items
        if items and self.current_history_index is not None:
            self.history_combo.current(self.current_history_index)
    
    def _load_history_selection(self):
        if not self.history:
            messagebox.showinfo("Info", "Aucun historique disponible")
            return
        
        idx = self.history_combo.current()
        if idx < 0 or idx >= len(self.history):
            messagebox.showwarning("Attention", "Sélectionnez une entrée")
            return
        
        entry = self.history[idx]
        self.current_history_index = idx
        
        rect = entry.get('rect')
        if rect and len(rect) == 4:
            self.selection_rect = tuple(rect)
            self.selected_components = self.parser.get_components_in_rect(*self.selection_rect)
        else:
            saved_refs = set(c.get('ref') for c in entry.get('components', []))
            self.selected_components = []
            for comp in self.parser.components:
                if comp.get('ref') in saved_refs:
                    bom_info = self.parser.get_bom_for_ref(comp['ref'], comp.get('id'))
                    self.selected_components.append({
                        'ref': comp['ref'],
                        'value': bom_info.get('value', ''),
                        'footprint': bom_info.get('footprint', ''),
                        'lcsc': bom_info.get('lcsc', ''),
                        'x': comp['x'],
                        'y': comp['y'],
                        'layer': comp['layer']
                    })
        
        self.processed_items.clear()
        for proc in entry.get('processed', []):
            if isinstance(proc, list) and len(proc) == 3:
                self.processed_items.add(tuple(proc))
        
        self._apply_filters()
        self._draw_main_pcb()
        
        self.export_btn.config(state=tk.NORMAL)
        self.export_csv_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)
        
        self.status_var.set(f"Chargé: {entry.get('name')} ({len(self.selected_components)} composants)")
    
    def _save_current_to_history(self):
        if not self.selected_components:
            messagebox.showwarning("Attention", "Aucune sélection à sauvegarder")
            return
        
        name = simpledialog.askstring("Nom", "Nom de la sélection:",
                                     initialvalue=f"Zone {len(self.history) + 1}")
        if not name:
            return
        
        entry = {
            'name': name,
            'date': datetime.now().strftime("%Y-%m-%d %H:%M"),
            'rect': list(self.selection_rect) if self.selection_rect else None,
            'components': [{'ref': c['ref'], 'value': c['value'],
                           'footprint': c['footprint'], 'lcsc': c['lcsc']}
                          for c in self.selected_components],
            'processed': [list(p) for p in self.processed_items]
        }
        
        self.history.append(entry)
        self.current_history_index = len(self.history) - 1
        self._save_history()
        self._update_history_combo()
        
        messagebox.showinfo("Succès", f"Sélection '{name}' sauvegardée")
    
    def _delete_history_selection(self):
        if not self.history:
            return
        
        idx = self.history_combo.current()
        if idx < 0 or idx >= len(self.history):
            return
        
        name = self.history[idx].get('name', f"Sélection {idx + 1}")
        if messagebox.askyesno("Confirmation", f"Supprimer '{name}'?"):
            del self.history[idx]
            self.current_history_index = None
            self._save_history()
            self._update_history_combo()
            self.history_var.set("")
    
    def _setup_keyboard_shortcuts(self):
        """Configure les raccourcis clavier"""
        self.root.bind('<Control-o>', lambda e: self._browse_file())
        self.root.bind('<Control-s>', lambda e: self._export_excel() if self.filtered_components else None)
        self.root.bind('<Control-l>', lambda e: self._load_file())
        self.root.bind('<Escape>', lambda e: self._clear_selection())
        self.root.bind('<Control-f>', lambda e: self.search_entry.focus_set())
        self.root.bind('<F5>', lambda e: self._draw_main_pcb() if self.parser else None)
        self.root.bind('<Left>', lambda e: self._navigate_prev())
        self.root.bind('<Right>', lambda e: self._navigate_next())
        self.root.bind('<space>', lambda e: self._toggle_processed())
    
    def _auto_load_bom(self):
        """Charge automatiquement bom/ibom.html"""
        script_dir = Path(__file__).parent
        bom_paths = [
            script_dir / 'bom' / 'ibom.html',
            script_dir / 'bom' / 'bom.html',
            Path('bom') / 'ibom.html',
        ]
        
        for bom_path in bom_paths:
            if bom_path.exists():
                self.file_var.set(str(bom_path.resolve()))
                self.root.after(100, self._load_file)
                break
    
    def run(self):
        """Lance l'application"""
        self.root.mainloop()


if __name__ == '__main__':
    app = IBomSelectorApp()
    app.run()

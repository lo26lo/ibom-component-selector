"""
Programme de sélection de composants depuis un fichier InteractiveHtmlBom
- Charge un fichier HTML généré par InteractiveHtmlBom
- Permet de sélectionner une zone rectangulaire sur le PCB
- Exporte les composants sélectionnés vers Excel ou CSV
"""

import csv
import json
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("Installation de openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

try:
    from lzstring import LZString as LZStringLib
    HAS_LZSTRING = True
except ImportError:
    HAS_LZSTRING = False
    print("lzstring non disponible, utilisation du décompresseur intégré")


class LZString:
    """Décompresseur LZ-String pour les données InteractiveHtmlBom"""
    
    @staticmethod
    def _get_base_value(char, alphabet):
        return alphabet.get(char, -1)
    
    @staticmethod
    def decompress_from_base64(compressed):
        """Décompresse une chaîne encodée en base64"""
        if not compressed:
            return ""
        
        # Caractères base64 standard
        key_str = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/="
        base_reverse_dict = {char: i for i, char in enumerate(key_str)}
        
        # Convertir base64 en données binaires
        try:
            length = len(compressed)
            get_next_value = lambda index: base_reverse_dict.get(compressed[index], 0)
            
            result = LZString._decompress(length, 32, get_next_value)
            return result
        except Exception as e:
            print(f"Erreur de décompression: {e}")
            return None
    
    @staticmethod
    def _decompress(length, reset_value, get_next_value):
        """Algorithme de décompression LZ"""
        dictionary = {}
        enlargeIn = 4
        dictSize = 4
        numBits = 3
        entry = ""
        result = []
        
        data_val = get_next_value(0)
        data_position = reset_value
        data_index = 1
        
        # Initialiser le dictionnaire
        for i in range(3):
            dictionary[i] = i
        
        bits = 0
        maxpower = 2 ** 2
        power = 1
        
        while power != maxpower:
            resb = data_val & data_position
            data_position >>= 1
            if data_position == 0:
                data_position = reset_value
                data_val = get_next_value(data_index)
                data_index += 1
            bits |= (1 if resb > 0 else 0) * power
            power <<= 1
        
        next_val = bits
        if next_val == 0:
            bits = 0
            maxpower = 2 ** 8
            power = 1
            while power != maxpower:
                resb = data_val & data_position
                data_position >>= 1
                if data_position == 0:
                    data_position = reset_value
                    data_val = get_next_value(data_index)
                    data_index += 1
                bits |= (1 if resb > 0 else 0) * power
                power <<= 1
            c = chr(bits)
        elif next_val == 1:
            bits = 0
            maxpower = 2 ** 16
            power = 1
            while power != maxpower:
                resb = data_val & data_position
                data_position >>= 1
                if data_position == 0:
                    data_position = reset_value
                    data_val = get_next_value(data_index)
                    data_index += 1
                bits |= (1 if resb > 0 else 0) * power
                power <<= 1
            c = chr(bits)
        elif next_val == 2:
            return ""
        
        dictionary[3] = c
        w = c
        result.append(c)
        
        while True:
            if data_index > length:
                return ""
            
            bits = 0
            maxpower = 2 ** numBits
            power = 1
            while power != maxpower:
                resb = data_val & data_position
                data_position >>= 1
                if data_position == 0:
                    data_position = reset_value
                    data_val = get_next_value(data_index)
                    data_index += 1
                bits |= (1 if resb > 0 else 0) * power
                power <<= 1
            
            c = bits
            if c == 0:
                bits = 0
                maxpower = 2 ** 8
                power = 1
                while power != maxpower:
                    resb = data_val & data_position
                    data_position >>= 1
                    if data_position == 0:
                        data_position = reset_value
                        data_val = get_next_value(data_index)
                        data_index += 1
                    bits |= (1 if resb > 0 else 0) * power
                    power <<= 1
                dictionary[dictSize] = chr(bits)
                dictSize += 1
                c = dictSize - 1
                enlargeIn -= 1
            elif c == 1:
                bits = 0
                maxpower = 2 ** 16
                power = 1
                while power != maxpower:
                    resb = data_val & data_position
                    data_position >>= 1
                    if data_position == 0:
                        data_position = reset_value
                        data_val = get_next_value(data_index)
                        data_index += 1
                    bits |= (1 if resb > 0 else 0) * power
                    power <<= 1
                dictionary[dictSize] = chr(bits)
                dictSize += 1
                c = dictSize - 1
                enlargeIn -= 1
            elif c == 2:
                return "".join(result)
            
            if enlargeIn == 0:
                enlargeIn = 2 ** numBits
                numBits += 1
            
            if c in dictionary:
                entry = dictionary[c] if isinstance(dictionary[c], str) else chr(dictionary[c])
            else:
                if c == dictSize:
                    entry = w + w[0]
                else:
                    return None
            
            result.append(entry)
            dictionary[dictSize] = w + entry[0]
            dictSize += 1
            enlargeIn -= 1
            
            if enlargeIn == 0:
                enlargeIn = 2 ** numBits
                numBits += 1
            
            w = entry
        
        return "".join(result)


class IBomParser:
    """Parse le fichier HTML d'InteractiveHtmlBom pour extraire les données"""
    
    def __init__(self, html_file_path):
        self.html_file_path = html_file_path
        self.pcbdata = None
        self.config = None
        self.components = []
        self.bom_data = []
        self.board_bbox = None
        self.lcsc_data = {}  # Mapping ref -> LCSC code
        
    def _load_lcsc_csv(self):
        """Charge le fichier CSV LCSC s'il existe"""
        # Chercher le fichier CSV dans différents emplacements
        html_dir = Path(self.html_file_path).parent
        possible_paths = [
            html_dir.parent / 'lcsc' / 'BOM-lcsc.csv',
            html_dir / 'lcsc' / 'BOM-lcsc.csv',
            Path('lcsc') / 'BOM-lcsc.csv',
            html_dir.parent / 'BOM-lcsc.csv',
        ]
        
        csv_path = None
        for path in possible_paths:
            if path.exists():
                csv_path = path
                break
        
        if not csv_path:
            return
        
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    designators = row.get('Designator', '')
                    lcsc_code = row.get('LCSC', '').strip()
                    
                    if lcsc_code:
                        # Les designators peuvent être multiples: "C1,C2,C3"
                        for ref in designators.split(','):
                            ref = ref.strip()
                            if ref:
                                self.lcsc_data[ref] = lcsc_code
            
            print(f"Fichier LCSC chargé: {len(self.lcsc_data)} références")
        except Exception as e:
            print(f"Erreur lors du chargement du fichier LCSC: {e}")
        
    def parse(self):
        """Parse le fichier HTML et extrait les données PCB"""
        with open(self.html_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Chercher les données compressées (nouveau format avec LZ-String)
        # Pattern: var pcbdata = JSON.parse(LZString.decompressFromBase64("..."))
        lz_match = re.search(r'LZString\.decompressFromBase64\(["\']([^"\']+)["\']\)', content)
        
        if lz_match:
            # Données compressées
            compressed_data = lz_match.group(1)
            print(f"Données compressées trouvées ({len(compressed_data)} caractères)")
            
            try:
                # Utiliser la bibliothèque lzstring si disponible
                if HAS_LZSTRING:
                    lz = LZStringLib()
                    decompressed = lz.decompressFromBase64(compressed_data)
                else:
                    decompressed = LZString.decompress_from_base64(compressed_data)
                
                if decompressed:
                    self.pcbdata = json.loads(decompressed)
                    print(f"Décompression réussie!")
                else:
                    raise ValueError("Échec de la décompression LZ-String")
            except Exception as e:
                print(f"Erreur lors de la décompression: {e}")
                raise ValueError(f"Impossible de décompresser les données: {e}")
        else:
            # Essayer le format non compressé (ancien format)
            pcbdata_match = re.search(r'var\s+pcbdata\s*=\s*(\{.*?\});', content, re.DOTALL)
            if not pcbdata_match:
                pcbdata_match = re.search(r'pcbdata\s*=\s*(\{.*?\})\s*[;\n]', content, re.DOTALL)
            
            if pcbdata_match:
                pcbdata_str = pcbdata_match.group(1)
                try:
                    self.pcbdata = json.loads(pcbdata_str)
                except json.JSONDecodeError:
                    pcbdata_str = re.sub(r',\s*}', '}', pcbdata_str)
                    pcbdata_str = re.sub(r',\s*]', ']', pcbdata_str)
                    self.pcbdata = json.loads(pcbdata_str)
            else:
                raise ValueError("Impossible de trouver les données pcbdata dans le fichier HTML")
        
        # Charger le fichier LCSC CSV
        self._load_lcsc_csv()
        
        self._extract_components()
        self._extract_bom()
        self._calculate_board_bbox()
        
        return self
    
    def _extract_components(self):
        """Extrait les composants avec leurs positions"""
        self.components = []
        
        if 'modules' in self.pcbdata:
            # Format ancien
            modules = self.pcbdata.get('modules', {})
            for layer in ['F', 'B']:
                layer_modules = modules.get(layer, [])
                for module in layer_modules:
                    self._add_component_from_module(module, layer)
        
        if 'footprints' in self.pcbdata:
            # Format nouveau - l'index correspond à l'ID du composant
            footprints = self.pcbdata.get('footprints', [])
            for fp_id, fp in enumerate(footprints):
                self._add_component_from_footprint(fp, fp_id)
    
    def _add_component_from_module(self, module, layer):
        """Ajoute un composant depuis un module (ancien format)"""
        ref = module.get('ref', '')
        bbox = module.get('bbox', {})
        
        # Calculer le centre du composant
        if bbox:
            x = (bbox.get('minx', 0) + bbox.get('maxx', 0)) / 2
            y = (bbox.get('miny', 0) + bbox.get('maxy', 0)) / 2
        else:
            x = module.get('x', 0)
            y = module.get('y', 0)
        
        self.components.append({
            'ref': ref,
            'x': x,
            'y': y,
            'layer': layer,
            'bbox': bbox
        })
    
    def _add_component_from_footprint(self, fp, fp_id):
        """Ajoute un composant depuis un footprint (nouveau format)"""
        ref = fp.get('ref', '')
        layer = fp.get('layer', 'F')
        
        # Position du composant - bbox.pos est le plus fiable
        x, y = 0, 0
        bbox = fp.get('bbox', {})
        
        # Priorité 1: bbox.pos (format le plus courant)
        if bbox and 'pos' in bbox:
            pos = bbox.get('pos', [0, 0])
            if isinstance(pos, list) and len(pos) >= 2:
                x, y = pos[0], pos[1]
        # Priorité 2: center (si présent dans certains formats)
        elif 'center' in fp:
            center = fp.get('center', [0, 0])
            if isinstance(center, list) and len(center) >= 2:
                x, y = center[0], center[1]
        # Priorité 3: ancien format de bbox avec minx/maxx
        elif bbox and 'minx' in bbox:
            x = (bbox.get('minx', 0) + bbox.get('maxx', 0)) / 2
            y = (bbox.get('miny', 0) + bbox.get('maxy', 0)) / 2
        # Priorité 4: moyenne des positions des pads
        else:
            pads = fp.get('pads', [])
            if pads:
                x = sum(p.get('pos', [0, 0])[0] for p in pads) / len(pads)
                y = sum(p.get('pos', [0, 0])[1] for p in pads) / len(pads)
        
        self.components.append({
            'ref': ref,
            'id': fp_id,
            'x': x,
            'y': y,
            'layer': layer,
            'bbox': bbox
        })
    
    def _extract_bom(self):
        """Extrait les données BOM selon le nouveau format"""
        self.bom_data = []
        
        bom = self.pcbdata.get('bom', {})
        fields_data = bom.get('fields', {})
        footprints = self.pcbdata.get('footprints', [])
        
        # Format both: liste de groupes de composants identiques
        # Chaque groupe: [[ref1, fp_id1], [ref2, fp_id2], ...]
        both = bom.get('both', [])
        
        for group in both:
            if not isinstance(group, list):
                continue
                
            for ref_item in group:
                if not isinstance(ref_item, list) or len(ref_item) < 2:
                    continue
                    
                ref_name = ref_item[0]
                fp_id = ref_item[1]
                
                # Récupérer value et footprint depuis fields_data
                value = ''
                footprint_name = ''
                lcsc = ''
                
                fp_id_str = str(fp_id)
                if fp_id_str in fields_data:
                    component_fields = fields_data[fp_id_str]
                    # fields_data[id] = [value, footprint_name, ...extra_fields...]
                    if len(component_fields) >= 1:
                        value = component_fields[0] or ''
                    if len(component_fields) >= 2:
                        footprint_name = component_fields[1] or ''
                    # Chercher un champ LCSC (commence par C suivi de chiffres)
                    for i, field_val in enumerate(component_fields):
                        if isinstance(field_val, str) and len(field_val) > 1:
                            # Pattern LCSC: C + chiffres (ex: C123456)
                            if field_val.startswith('C') and field_val[1:].isdigit():
                                lcsc = field_val
                                break
                
                # Si pas de LCSC trouvé dans le BOM, chercher dans le fichier CSV
                if not lcsc and ref_name in self.lcsc_data:
                    lcsc = self.lcsc_data[ref_name]
                
                self.bom_data.append({
                    'ref': ref_name,
                    'id': fp_id,
                    'value': value,
                    'footprint': footprint_name,
                    'lcsc': lcsc
                })
    
    def _calculate_board_bbox(self):
        """Calcule la bounding box du PCB"""
        edges = self.pcbdata.get('edges_bbox', {})
        if edges:
            self.board_bbox = {
                'minx': edges.get('minx', 0),
                'miny': edges.get('miny', 0),
                'maxx': edges.get('maxx', 100),
                'maxy': edges.get('maxy', 100)
            }
        else:
            # Calculer depuis les composants
            if self.components:
                xs = [c['x'] for c in self.components]
                ys = [c['y'] for c in self.components]
                margin = 5
                self.board_bbox = {
                    'minx': min(xs) - margin,
                    'miny': min(ys) - margin,
                    'maxx': max(xs) + margin,
                    'maxy': max(ys) + margin
                }
            else:
                self.board_bbox = {'minx': 0, 'miny': 0, 'maxx': 100, 'maxy': 100}
    
    def get_bom_for_ref(self, ref, fp_id=None):
        """Récupère les infos BOM pour une référence"""
        for bom_entry in self.bom_data:
            if bom_entry['ref'] == ref:
                return bom_entry
            if fp_id is not None and bom_entry.get('id') == fp_id:
                return bom_entry
        return {'ref': ref, 'value': '', 'footprint': '', 'lcsc': '', 'id': fp_id}
    
    def get_components_in_rect(self, x1, y1, x2, y2):
        """Retourne les composants dans le rectangle spécifié"""
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
        
        selected = []
        for comp in self.components:
            if min_x <= comp['x'] <= max_x and min_y <= comp['y'] <= max_y:
                bom_info = self.get_bom_for_ref(comp['ref'], comp.get('id'))
                selected.append({
                    'ref': comp['ref'],
                    'value': bom_info.get('value', ''),
                    'footprint': bom_info.get('footprint', ''),
                    'lcsc': bom_info.get('lcsc', ''),
                    'x': comp['x'],
                    'y': comp['y'],
                    'layer': comp['layer']
                })
        
        return selected


class PCBViewer(tk.Toplevel):
    """Fenêtre de visualisation du PCB avec sélection rectangulaire"""
    
    def __init__(self, parent, parser, callback):
        super().__init__(parent)
        self.parser = parser
        self.callback = callback
        
        self.title("Sélection de zone - PCB Viewer")
        self.geometry("900x700")
        
        # Variables de sélection
        self.start_x = None
        self.start_y = None
        self.rect_id = None
        self.scale = 1.0
        self.offset_x = 50
        self.offset_y = 50
        
        self._setup_ui()
        self._draw_pcb()
    
    def _setup_ui(self):
        """Configure l'interface utilisateur"""
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Instructions
        instructions = ttk.Label(
            main_frame, 
            text="Cliquez et glissez pour sélectionner une zone rectangulaire sur le PCB",
            font=('Segoe UI', 10)
        )
        instructions.pack(pady=(0, 10))
        
        # Canvas pour le PCB
        canvas_frame = ttk.Frame(main_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas = tk.Canvas(canvas_frame, bg='#1a1a2e', highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        h_scroll = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        v_scroll = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        
        self.canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        
        # Bindings
        self.canvas.bind('<Button-1>', self._on_mouse_down)
        self.canvas.bind('<B1-Motion>', self._on_mouse_drag)
        self.canvas.bind('<ButtonRelease-1>', self._on_mouse_up)
        self.canvas.bind('<MouseWheel>', self._on_mousewheel)
        
        # Boutons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, text="Zoom +", command=self._zoom_in).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Zoom -", command=self._zoom_out).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Réinitialiser", command=self._reset_view).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Annuler", command=self.destroy).pack(side=tk.LEFT, padx=5)
    
    def _pcb_to_canvas(self, x, y):
        """Convertit les coordonnées PCB en coordonnées canvas"""
        bbox = self.parser.board_bbox
        canvas_x = self.offset_x + (x - bbox['minx']) * self.scale
        canvas_y = self.offset_y + (y - bbox['miny']) * self.scale
        return canvas_x, canvas_y
    
    def _canvas_to_pcb(self, canvas_x, canvas_y):
        """Convertit les coordonnées canvas en coordonnées PCB"""
        bbox = self.parser.board_bbox
        x = (canvas_x - self.offset_x) / self.scale + bbox['minx']
        y = (canvas_y - self.offset_y) / self.scale + bbox['miny']
        return x, y
    
    def _draw_pcb(self, recalculate_scale=True):
        """Dessine le PCB et les composants avec détails"""
        self.canvas.delete('all')
        
        bbox = self.parser.board_bbox
        width = bbox['maxx'] - bbox['minx']
        height = bbox['maxy'] - bbox['miny']
        
        # Calculer l'échelle pour tenir dans le canvas (seulement si demandé)
        canvas_width = self.canvas.winfo_width() or 800
        canvas_height = self.canvas.winfo_height() or 600
        
        if recalculate_scale:
            scale_x = (canvas_width - 100) / width if width > 0 else 1
            scale_y = (canvas_height - 100) / height if height > 0 else 1
            self.scale = min(scale_x, scale_y) * 0.9
        
        # Fond du PCB
        x1, y1 = self._pcb_to_canvas(bbox['minx'], bbox['miny'])
        x2, y2 = self._pcb_to_canvas(bbox['maxx'], bbox['maxy'])
        self.canvas.create_rectangle(x1, y1, x2, y2, outline='', fill='#1a1a2e')
        
        # Dessiner les edges (contour du PCB)
        self._draw_edges()
        
        # Dessiner le silkscreen
        self._draw_silkscreen()
        
        # Dessiner les pads des footprints
        self._draw_pads()
        
        # Dessiner les références des composants
        self._draw_references()
        
        # Légende
        self.canvas.create_text(10, 10, anchor='nw', text="● Front", fill='#00ff88', font=('Segoe UI', 9))
        self.canvas.create_text(10, 25, anchor='nw', text="● Back", fill='#ff6b6b', font=('Segoe UI', 9))
    
    def _draw_edges(self):
        """Dessine le contour du PCB"""
        import math
        bbox = self.parser.board_bbox
        edges = self.parser.pcbdata.get('edges', [])
        
        for edge in edges:
            edge_type = edge.get('type', '')
            if edge_type == 'segment':
                start = edge.get('start', [0, 0])
                end = edge.get('end', [0, 0])
                
                # Ignorer les segments hors limites du PCB
                if (start[0] < bbox['minx'] - 5 or start[0] > bbox['maxx'] + 5 or
                    start[1] < bbox['miny'] - 5 or start[1] > bbox['maxy'] + 5 or
                    end[0] < bbox['minx'] - 5 or end[0] > bbox['maxx'] + 5 or
                    end[1] < bbox['miny'] - 5 or end[1] > bbox['maxy'] + 5):
                    continue
                
                x1, y1 = self._pcb_to_canvas(start[0], start[1])
                x2, y2 = self._pcb_to_canvas(end[0], end[1])
                width = max(1, edge.get('width', 0.15) * self.scale)
                self.canvas.create_line(x1, y1, x2, y2, fill='#ffffff', width=width)
            
            elif edge_type == 'arc':
                # Dessiner l'arc comme une série de segments
                start = edge.get('start', [0, 0])
                radius = edge.get('radius', 1)
                start_angle = edge.get('startangle', 0)
                end_angle = edge.get('endangle', 360)
                
                # Calculer le centre de l'arc
                cx = start[0]
                cy = start[1]
                
                # Créer les points de l'arc
                points = []
                num_segments = 20
                for i in range(num_segments + 1):
                    angle = math.radians(start_angle + (end_angle - start_angle) * i / num_segments)
                    px = cx + radius * math.cos(angle)
                    py = cy + radius * math.sin(angle)
                    canvas_x, canvas_y = self._pcb_to_canvas(px, py)
                    points.extend([canvas_x, canvas_y])
                
                if len(points) >= 4:
                    self.canvas.create_line(points, fill='#ffffff', width=1, smooth=True)
            
            elif edge_type == 'polygon':
                polygons = edge.get('polygons', [])
                for poly in polygons:
                    if len(poly) >= 3:
                        points = []
                        for pt in poly:
                            cx, cy = self._pcb_to_canvas(pt[0], pt[1])
                            points.extend([cx, cy])
                        if len(points) >= 6:
                            self.canvas.create_polygon(points, outline='#ffffff', fill='', width=1)
    
    def _draw_silkscreen(self):
        """Dessine le silkscreen (couche F seulement pour simplifier)"""
        drawings = self.parser.pcbdata.get('drawings', {})
        silkscreen = drawings.get('silkscreen', {})
        
        # Dessiner Front silkscreen
        front_silk = silkscreen.get('F', [])
        for drawing in front_silk:
            self._draw_silkscreen_element(drawing, '#cccccc')
        
        # Dessiner Back silkscreen en plus sombre
        back_silk = silkscreen.get('B', [])
        for drawing in back_silk:
            self._draw_silkscreen_element(drawing, '#555555')
    
    def _draw_silkscreen_element(self, drawing, color):
        """Dessine un élément de silkscreen"""
        draw_type = drawing.get('type', '')
        
        if draw_type == 'segment':
            start = drawing.get('start', [0, 0])
            end = drawing.get('end', [0, 0])
            x1, y1 = self._pcb_to_canvas(start[0], start[1])
            x2, y2 = self._pcb_to_canvas(end[0], end[1])
            self.canvas.create_line(x1, y1, x2, y2, fill=color, width=1)
        
        elif draw_type == 'polygon':
            polygons = drawing.get('polygons', [])
            for poly in polygons:
                if len(poly) >= 3:
                    points = []
                    for pt in poly:
                        cx, cy = self._pcb_to_canvas(pt[0], pt[1])
                        points.extend([cx, cy])
                    if len(points) >= 6:
                        self.canvas.create_polygon(points, outline=color, fill='', width=1)
        
        elif draw_type == 'circle':
            start = drawing.get('start', [0, 0])
            radius = drawing.get('radius', 1)
            cx, cy = self._pcb_to_canvas(start[0], start[1])
            r = radius * self.scale
            self.canvas.create_oval(cx - r, cy - r, cx + r, cy + r, outline=color, width=1)
        
        elif draw_type == 'arc':
            start = drawing.get('start', [0, 0])
            end = drawing.get('end', [0, 0])
            x1, y1 = self._pcb_to_canvas(start[0], start[1])
            x2, y2 = self._pcb_to_canvas(end[0], end[1])
            self.canvas.create_line(x1, y1, x2, y2, fill=color, width=1)
    
    def _draw_pads(self):
        """Dessine les pads de tous les footprints"""
        footprints = self.parser.pcbdata.get('footprints', [])
        
        for fp in footprints:
            layer = fp.get('layer', 'F')
            pads = fp.get('pads', [])
            
            for pad in pads:
                self._draw_pad(pad, layer)
    
    def _draw_pad(self, pad, fp_layer):
        """Dessine un pad individuel"""
        pos = pad.get('pos', [0, 0])
        size = pad.get('size', [1, 1])
        shape = pad.get('shape', 'rect')
        pad_type = pad.get('type', 'smd')
        layers = pad.get('layers', [fp_layer])
        
        cx, cy = self._pcb_to_canvas(pos[0], pos[1])
        
        # Taille du pad en pixels
        w = size[0] * self.scale
        h = size[1] * self.scale
        
        # Couleur selon le type et le layer
        if pad_type == 'th':  # Through-hole
            color = '#b8860b'  # Or foncé
            outline_color = '#ffd700'
        elif 'F' in layers:
            color = '#8b0000'  # Rouge foncé pour Front
            outline_color = '#ff4444'
        else:
            color = '#00008b'  # Bleu foncé pour Back
            outline_color = '#4444ff'
        
        # Dessiner selon la forme
        if shape == 'circle' or shape == 'oval':
            r = max(w, h) / 2
            self.canvas.create_oval(
                cx - r, cy - r, cx + r, cy + r,
                fill=color, outline=outline_color, width=1
            )
        elif shape == 'roundrect':
            # Approximer par un rectangle
            self.canvas.create_rectangle(
                cx - w/2, cy - h/2, cx + w/2, cy + h/2,
                fill=color, outline=outline_color, width=1
            )
        else:  # rect ou autre
            self.canvas.create_rectangle(
                cx - w/2, cy - h/2, cx + w/2, cy + h/2,
                fill=color, outline=outline_color, width=1
            )
    
    def _draw_references(self):
        """Dessine les références des composants"""
        # Afficher les références seulement si zoom suffisant
        if self.scale < 2:
            return
        
        for comp in self.parser.components:
            ref = comp.get('ref', '')
            if not ref or ref == 'REF**':
                continue
            
            cx, cy = self._pcb_to_canvas(comp['x'], comp['y'])
            color = '#00ff88' if comp['layer'] == 'F' else '#ff6b6b'
            
            # Taille de police adaptative
            font_size = max(6, min(10, int(self.scale * 1.5)))
            
            self.canvas.create_text(
                cx, cy,
                text=ref,
                fill=color,
                font=('Consolas', font_size, 'bold')
            )
    
    def _on_mouse_down(self, event):
        """Début de la sélection"""
        self.start_x = event.x
        self.start_y = event.y
        if self.rect_id:
            self.canvas.delete(self.rect_id)
    
    def _on_mouse_drag(self, event):
        """Pendant la sélection"""
        if self.rect_id:
            self.canvas.delete(self.rect_id)
        
        self.rect_id = self.canvas.create_rectangle(
            self.start_x, self.start_y, event.x, event.y,
            outline='#ffcc00', width=2, dash=(5, 5)
        )
    
    def _on_mouse_up(self, event):
        """Fin de la sélection"""
        if self.start_x is None:
            return
        
        # Convertir en coordonnées PCB
        pcb_x1, pcb_y1 = self._canvas_to_pcb(self.start_x, self.start_y)
        pcb_x2, pcb_y2 = self._canvas_to_pcb(event.x, event.y)
        
        # Récupérer les composants sélectionnés
        selected = self.parser.get_components_in_rect(pcb_x1, pcb_y1, pcb_x2, pcb_y2)
        
        if selected:
            # Normaliser les coordonnées du rectangle
            rect = (min(pcb_x1, pcb_x2), min(pcb_y1, pcb_y2), 
                    max(pcb_x1, pcb_x2), max(pcb_y1, pcb_y2))
            self.callback(selected, rect)
            self.destroy()
        else:
            messagebox.showinfo("Sélection", "Aucun composant dans la zone sélectionnée.\nEssayez une autre zone.")
            if self.rect_id:
                self.canvas.delete(self.rect_id)
    
    def _on_mousewheel(self, event):
        """Zoom avec la molette"""
        if event.delta > 0:
            self._zoom_in()
        else:
            self._zoom_out()
    
    def _zoom_in(self):
        self.scale *= 1.2
        self._draw_pcb(recalculate_scale=False)

    def _zoom_out(self):
        self.scale /= 1.2
        self._draw_pcb(recalculate_scale=False)

    def _reset_view(self):
        self._draw_pcb(recalculate_scale=True)
class IBomSelectorApp:
    """Application principale"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("IBom Component Selector")
        self.root.geometry("1100x900")
        self.root.configure(bg='#f0f0f0')
        
        self.parser = None
        self.selected_components = []
        self.filtered_components = []  # Composants filtrés pour l'affichage
        self.selection_rect = None  # Rectangle de sélection en coordonnées PCB
        self.layer_filter = tk.StringVar(value="all")  # Filtre de couche: all, F, B
        self.search_var = tk.StringVar()  # Variable de recherche
        self.processed_items = set()  # Ensemble des items traités (clés: value, footprint, lcsc)
        self.sort_column = None  # Colonne de tri actuelle
        self.sort_reverse = False  # Ordre de tri inversé
        self.history = []  # Historique des sélections
        self.history_file = None  # Fichier d'historique associé au HTML courant
        self.current_history_index = None  # Index de l'élément d'historique actuel
        
        self._setup_ui()
        self._setup_keyboard_shortcuts()
    
    def _setup_ui(self):
        """Configure l'interface utilisateur principale"""
        # Style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Titre
        title = ttk.Label(
            main_frame,
            text="InteractiveHtmlBom - Sélecteur de Composants",
            font=('Segoe UI', 16, 'bold')
        )
        title.pack(pady=(0, 20))
        
        # Frame pour le fichier
        file_frame = ttk.LabelFrame(main_frame, text="Fichier HTML", padding=10)
        file_frame.pack(fill=tk.X, pady=10)
        
        self.file_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_var, width=60)
        file_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        browse_btn = ttk.Button(file_frame, text="Parcourir...", command=self._browse_file)
        browse_btn.pack(side=tk.LEFT)
        
        load_btn = ttk.Button(file_frame, text="Charger", command=self._load_file)
        load_btn.pack(side=tk.LEFT, padx=5)
        
        # Frame pour la miniature du PCB et les boutons
        pcb_frame = ttk.LabelFrame(main_frame, text="PCB - Cliquez pour sélectionner une zone", padding=5)
        pcb_frame.pack(fill=tk.X, pady=10)
        
        # Frame interne pour centrer le contenu
        pcb_inner_frame = ttk.Frame(pcb_frame)
        pcb_inner_frame.pack(anchor='center')
        
        # Canvas pour la miniature du PCB (à gauche)
        self.pcb_canvas = tk.Canvas(pcb_inner_frame, width=700, height=180, bg='#1a1a2e', 
                                     highlightthickness=1, highlightbackground='#4a4a6a')
        self.pcb_canvas.pack(side=tk.LEFT, pady=5, padx=5)
        self.pcb_canvas.bind('<Button-1>', self._on_pcb_click)
        
        # Frame pour les boutons (à droite du PCB)
        btn_side_frame = ttk.Frame(pcb_inner_frame)
        btn_side_frame.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=10)
        
        self.clear_btn = ttk.Button(
            btn_side_frame,
            text="Effacer la sélection",
            command=self._clear_selection,
            state=tk.DISABLED,
            width=20
        )
        self.clear_btn.pack(pady=10)
        
        self.export_btn = ttk.Button(
            btn_side_frame,
            text="Exporter vers Excel",
            command=self._export_excel,
            state=tk.DISABLED,
            width=20
        )
        self.export_btn.pack(pady=10)
        
        self.export_csv_btn = ttk.Button(
            btn_side_frame,
            text="Exporter vers CSV",
            command=self._export_csv,
            state=tk.DISABLED,
            width=20
        )
        self.export_csv_btn.pack(pady=10)
        
        # Label de statut
        self.status_var = tk.StringVar(value="Chargez un fichier HTML InteractiveHtmlBom pour commencer")
        status_label = ttk.Label(btn_side_frame, textvariable=self.status_var, font=('Segoe UI', 9), wraplength=150)
        status_label.pack(pady=20)
        
        # Frame pour l'historique
        history_frame = ttk.LabelFrame(main_frame, text="Historique des sélections", padding=10)
        history_frame.pack(fill=tk.X, pady=5)
        
        # Combobox pour l'historique
        history_inner = ttk.Frame(history_frame)
        history_inner.pack(fill=tk.X)
        
        ttk.Label(history_inner, text="Sélection:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.history_var = tk.StringVar()
        self.history_combo = ttk.Combobox(history_inner, textvariable=self.history_var, 
                                           state='readonly', width=50)
        self.history_combo.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.history_combo.bind('<<ComboboxSelected>>', self._on_history_select)
        
        ttk.Button(history_inner, text="Charger", command=self._load_history_selection,
                   width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(history_inner, text="Sauvegarder", command=self._save_current_to_history,
                   width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(history_inner, text="Supprimer", command=self._delete_history_selection,
                   width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(history_inner, text="Mettre à jour", command=self._update_history_selection,
                   width=12).pack(side=tk.LEFT, padx=2)
        
        # Message initial sur le canvas
        self.pcb_canvas.create_text(350, 90, text="Chargez un fichier pour voir le PCB", 
                                     fill='#666666', font=('Segoe UI', 12))
        
        # Frame pour les filtres
        filter_frame = ttk.LabelFrame(main_frame, text="Filtres", padding=10)
        filter_frame.pack(fill=tk.X, pady=5)
        
        # Filtre par couche
        layer_label = ttk.Label(filter_frame, text="Couche:")
        layer_label.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Radiobutton(filter_frame, text="Toutes", variable=self.layer_filter, 
                        value="all", command=self._apply_filters).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(filter_frame, text="Front (F)", variable=self.layer_filter, 
                        value="F", command=self._apply_filters).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(filter_frame, text="Back (B)", variable=self.layer_filter, 
                        value="B", command=self._apply_filters).pack(side=tk.LEFT, padx=5)
        
        # Séparateur
        ttk.Separator(filter_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        # Recherche
        search_label = ttk.Label(filter_frame, text="Rechercher:")
        search_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_var.trace('w', lambda *args: self._apply_filters())
        
        clear_search_btn = ttk.Button(filter_frame, text="✕", width=3, command=self._clear_search)
        clear_search_btn.pack(side=tk.LEFT, padx=2)
        
        # Statistiques
        self.stats_var = tk.StringVar(value="")
        stats_label = ttk.Label(filter_frame, textvariable=self.stats_var, font=('Segoe UI', 9, 'italic'))
        stats_label.pack(side=tk.RIGHT, padx=10)
        
        # Liste des composants sélectionnés
        list_frame = ttk.LabelFrame(main_frame, text="Composants sélectionnés (cliquez sur les en-têtes pour trier)", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Treeview pour afficher les composants
        columns = ('done', 'qty', 'ref', 'value', 'footprint', 'lcsc')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        # Configuration des tags pour le style
        self.tree.tag_configure('done', background='#c8e6c9', foreground='#2e7d32')
        self.tree.tag_configure('pending', background='', foreground='')
        
        # En-têtes avec tri
        self.tree.heading('done', text='✓', command=lambda: self._sort_by_column('done'))
        self.tree.heading('qty', text='Qté ↕', command=lambda: self._sort_by_column('qty'))
        self.tree.heading('ref', text='Références ↕', command=lambda: self._sort_by_column('ref'))
        self.tree.heading('value', text='Valeur ↕', command=lambda: self._sort_by_column('value'))
        self.tree.heading('footprint', text='Footprint ↕', command=lambda: self._sort_by_column('footprint'))
        self.tree.heading('lcsc', text='LCSC ↕', command=lambda: self._sort_by_column('lcsc'))
        
        self.tree.column('done', width=30, anchor='center')
        self.tree.column('qty', width=50, anchor='center')
        self.tree.column('ref', width=150)
        self.tree.column('value', width=120)
        self.tree.column('footprint', width=180)
        self.tree.column('lcsc', width=100)
        
        # Bind double-clic pour marquer comme traité
        self.tree.bind('<Double-1>', self._toggle_processed)
        self.tree.bind('<space>', self._toggle_processed)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Boutons pour la gestion des "traités"
        processed_frame = ttk.Frame(main_frame)
        processed_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(processed_frame, text="Marquer sélection comme traitée", 
                   command=self._mark_selected_processed).pack(side=tk.LEFT, padx=5)
        ttk.Button(processed_frame, text="Démarquer sélection", 
                   command=self._unmark_selected_processed).pack(side=tk.LEFT, padx=5)
        ttk.Button(processed_frame, text="Tout démarquer", 
                   command=self._unmark_all_processed).pack(side=tk.LEFT, padx=5)
        
        self.processed_count_var = tk.StringVar(value="")
        ttk.Label(processed_frame, textvariable=self.processed_count_var, 
                  font=('Segoe UI', 9, 'italic')).pack(side=tk.RIGHT, padx=10)
    
    def _browse_file(self):
        """Ouvre le dialogue de sélection de fichier"""
        filename = filedialog.askopenfilename(
            title="Sélectionner un fichier InteractiveHtmlBom",
            filetypes=[("Fichiers HTML", "*.html"), ("Tous les fichiers", "*.*")]
        )
        if filename:
            self.file_var.set(filename)
    
    def _load_file(self):
        """Charge et parse le fichier HTML"""
        filepath = self.file_var.get()
        if not filepath:
            messagebox.showwarning("Attention", "Veuillez sélectionner un fichier HTML")
            return
        
        if not Path(filepath).exists():
            messagebox.showerror("Erreur", "Le fichier n'existe pas")
            return
        
        try:
            self.parser = IBomParser(filepath)
            self.parser.parse()
            
            self.status_var.set(
                f"Fichier chargé: {len(self.parser.components)} composants trouvés"
            )
            
            # Charger l'historique associé au fichier
            self._load_history()
            
            # Dessiner la miniature du PCB
            self._draw_pcb_preview()
            
            messagebox.showinfo(
                "Succès",
                f"Fichier chargé avec succès!\n"
                f"Composants trouvés: {len(self.parser.components)}\n"
                f"Entrées BOM: {len(self.parser.bom_data)}\n"
                f"Sélections en historique: {len(self.history)}"
            )
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement:\n{str(e)}")
            self.status_var.set("Erreur lors du chargement du fichier")
    
    def _on_pcb_click(self, event):
        """Ouvre le viewer PCB quand on clique sur la miniature"""
        if self.parser:
            self._open_pcb_viewer()
    
    def _draw_pcb_preview(self):
        """Dessine une miniature du PCB dans le canvas principal"""
        self.pcb_canvas.delete('all')
        
        if not self.parser:
            return
        
        bbox = self.parser.board_bbox
        pcb_width = bbox['maxx'] - bbox['minx']
        pcb_height = bbox['maxy'] - bbox['miny']
        
        # Calculer l'échelle pour tenir dans le canvas
        canvas_width = self.pcb_canvas.winfo_width() or 700
        canvas_height = self.pcb_canvas.winfo_height() or 180
        
        scale_x = (canvas_width - 20) / pcb_width if pcb_width > 0 else 1
        scale_y = (canvas_height - 20) / pcb_height if pcb_height > 0 else 1
        self.preview_scale = min(scale_x, scale_y)
        
        # Offset pour centrer
        self.preview_offset_x = (canvas_width - pcb_width * self.preview_scale) / 2
        self.preview_offset_y = (canvas_height - pcb_height * self.preview_scale) / 2
        
        # Fond du PCB
        x1 = self.preview_offset_x
        y1 = self.preview_offset_y
        x2 = x1 + pcb_width * self.preview_scale
        y2 = y1 + pcb_height * self.preview_scale
        self.pcb_canvas.create_rectangle(x1, y1, x2, y2, outline='#ffffff', fill='#16213e', width=2)
        
        # Dessiner les pads (simplifiés)
        for fp in self.parser.pcbdata.get('footprints', []):
            layer = fp.get('layer', 'F')
            for pad in fp.get('pads', []):
                pos = pad.get('pos', [0, 0])
                size = pad.get('size', [0.5, 0.5])
                
                px = self.preview_offset_x + (pos[0] - bbox['minx']) * self.preview_scale
                py = self.preview_offset_y + (pos[1] - bbox['miny']) * self.preview_scale
                
                w = max(2, size[0] * self.preview_scale / 2)
                h = max(2, size[1] * self.preview_scale / 2)
                
                color = '#8b0000' if 'F' in pad.get('layers', [layer]) else '#00008b'
                self.pcb_canvas.create_rectangle(px - w, py - h, px + w, py + h, fill=color, outline='')
        
        # Dessiner la zone de sélection si elle existe
        if self.selection_rect:
            sx1, sy1, sx2, sy2 = self.selection_rect
            cx1 = self.preview_offset_x + (sx1 - bbox['minx']) * self.preview_scale
            cy1 = self.preview_offset_y + (sy1 - bbox['miny']) * self.preview_scale
            cx2 = self.preview_offset_x + (sx2 - bbox['minx']) * self.preview_scale
            cy2 = self.preview_offset_y + (sy2 - bbox['miny']) * self.preview_scale
            self.pcb_canvas.create_rectangle(cx1, cy1, cx2, cy2, 
                                              outline='#ffcc00', width=2, dash=(5, 3))
        
        # Texte indicatif
        self.pcb_canvas.create_text(canvas_width / 2, canvas_height - 10, 
                                     text="Cliquez pour sélectionner une zone", 
                                     fill='#888888', font=('Segoe UI', 9))
    
    def _open_pcb_viewer(self):
        """Ouvre la fenêtre de visualisation du PCB"""
        if not self.parser:
            return
        
        viewer = PCBViewer(self.root, self.parser, self._on_selection)
        viewer.transient(self.root)
        viewer.grab_set()
        
        # Attendre que la fenêtre soit affichée avant de dessiner
        viewer.update()
        viewer._draw_pcb()
    
    def _on_selection(self, selected_components, selection_rect=None):
        """Callback quand des composants sont sélectionnés"""
        self.selected_components = selected_components
        self.selection_rect = selection_rect
        
        self._apply_filters()
        self._draw_pcb_preview()  # Redessiner avec la zone de sélection
        
        self.export_btn.config(state=tk.NORMAL)
        self.export_csv_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)
        self.status_var.set(f"{len(selected_components)} composants sélectionnés")
    
    def _apply_filters(self):
        """Applique les filtres sur les composants sélectionnés"""
        layer_filter = self.layer_filter.get()
        search_text = self.search_var.get().lower().strip()
        
        # Filtrer les composants
        self.filtered_components = []
        for comp in self.selected_components:
            # Filtre par couche
            if layer_filter != "all" and comp.get('layer', 'F') != layer_filter:
                continue
            
            # Filtre par recherche
            if search_text:
                searchable = f"{comp['ref']} {comp['value']} {comp['footprint']} {comp['lcsc']}".lower()
                if search_text not in searchable:
                    continue
            
            self.filtered_components.append(comp)
        
        self._update_tree()
        self._update_statistics()
    
    def _update_statistics(self):
        """Met à jour les statistiques affichées"""
        if not self.selected_components:
            self.stats_var.set("")
            return
        
        total = len(self.selected_components)
        filtered = len(self.filtered_components)
        front_count = sum(1 for c in self.selected_components if c.get('layer', 'F') == 'F')
        back_count = total - front_count
        
        if filtered == total:
            self.stats_var.set(f"Total: {total} | Front: {front_count} | Back: {back_count}")
        else:
            self.stats_var.set(f"Affichés: {filtered}/{total} | Front: {front_count} | Back: {back_count}")
    
    def _clear_search(self):
        """Efface le champ de recherche"""
        self.search_var.set("")
    
    def _update_tree(self):
        """Met à jour l'affichage de la liste des composants (regroupés par value/footprint)"""
        # Effacer les anciens éléments
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Regrouper par (value, footprint, lcsc)
        grouped = {}
        for comp in self.filtered_components:
            key = (comp['value'], comp['footprint'], comp['lcsc'])
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(comp['ref'])
        
        # Construire la liste des données pour le tri
        data_list = []
        for (value, footprint, lcsc), refs in grouped.items():
            refs_sorted = sorted(refs, key=lambda r: (r[0], int(''.join(filter(str.isdigit, r)) or 0)))
            refs_str = ', '.join(refs_sorted)
            key = (value, footprint, lcsc)
            is_done = key in self.processed_items
            data_list.append({
                'key': key,
                'done': '✓' if is_done else '',
                'qty': len(refs),
                'ref': refs_str,
                'value': value,
                'footprint': footprint,
                'lcsc': lcsc,
                'is_done': is_done
            })
        
        # Appliquer le tri
        if self.sort_column:
            if self.sort_column == 'done':
                data_list.sort(key=lambda x: (not x['is_done'], x['value']), reverse=self.sort_reverse)
            elif self.sort_column == 'qty':
                data_list.sort(key=lambda x: x['qty'], reverse=self.sort_reverse)
            elif self.sort_column == 'ref':
                data_list.sort(key=lambda x: x['ref'], reverse=self.sort_reverse)
            elif self.sort_column == 'value':
                data_list.sort(key=lambda x: x['value'], reverse=self.sort_reverse)
            elif self.sort_column == 'footprint':
                data_list.sort(key=lambda x: x['footprint'], reverse=self.sort_reverse)
            elif self.sort_column == 'lcsc':
                data_list.sort(key=lambda x: x['lcsc'], reverse=self.sort_reverse)
        else:
            # Tri par défaut: valeur puis référence
            data_list.sort(key=lambda x: (x['value'], x['ref']))
        
        # Ajouter les groupes
        for data in data_list:
            tag = 'done' if data['is_done'] else 'pending'
            self.tree.insert('', tk.END, values=(
                data['done'],
                data['qty'],
                data['ref'],
                data['value'],
                data['footprint'],
                data['lcsc']
            ), tags=(tag,))
        
        self._update_processed_count()
    
    def _sort_by_column(self, column):
        """Trie le tableau par la colonne spécifiée"""
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        
        # Mettre à jour les indicateurs de tri dans les en-têtes
        for col in ('done', 'qty', 'ref', 'value', 'footprint', 'lcsc'):
            text = self.tree.heading(col)['text'].rstrip(' ↑↓↕')
            if col == column:
                arrow = ' ↓' if self.sort_reverse else ' ↑'
            else:
                arrow = ' ↕' if col != 'done' else ''
            self.tree.heading(col, text=text + arrow)
        
        self._update_tree()
    
    def _toggle_processed(self, event=None):
        """Bascule l'état 'traité' de la ligne sélectionnée"""
        selection = self.tree.selection()
        if not selection:
            return
        
        for item in selection:
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                # Récupérer la clé (value, footprint, lcsc)
                key = (values[3], values[4], values[5])  # value, footprint, lcsc
                
                if key in self.processed_items:
                    self.processed_items.discard(key)
                else:
                    self.processed_items.add(key)
        
        self._update_tree()
    
    def _mark_selected_processed(self):
        """Marque les lignes sélectionnées comme traitées"""
        selection = self.tree.selection()
        for item in selection:
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                self.processed_items.add(key)
        self._update_tree()
    
    def _unmark_selected_processed(self):
        """Démarque les lignes sélectionnées"""
        selection = self.tree.selection()
        for item in selection:
            values = self.tree.item(item, 'values')
            if len(values) >= 6:
                key = (values[3], values[4], values[5])
                self.processed_items.discard(key)
        self._update_tree()
    
    def _unmark_all_processed(self):
        """Démarque toutes les lignes"""
        self.processed_items.clear()
        self._update_tree()
    
    def _update_processed_count(self):
        """Met à jour le compteur d'items traités"""
        total = len(self.tree.get_children())
        processed = sum(1 for item in self.tree.get_children() 
                       if self.tree.item(item, 'values')[0] == '✓')
        if total > 0:
            self.processed_count_var.set(f"Traités: {processed}/{total}")
        else:
            self.processed_count_var.set("")
    
    def _clear_selection(self):
        """Efface la sélection actuelle"""
        self.selected_components = []
        self.filtered_components = []
        self.selection_rect = None
        self.processed_items.clear()  # Réinitialiser aussi les items traités
        self._update_tree()
        self._update_statistics()
        self._draw_pcb_preview()  # Redessiner sans la zone de sélection
        self.export_btn.config(state=tk.DISABLED)
        self.export_csv_btn.config(state=tk.DISABLED)
        self.clear_btn.config(state=tk.DISABLED)
        self.status_var.set("Sélection effacée")
    
    def _export_excel(self):
        """Exporte les composants sélectionnés vers Excel"""
        if not self.filtered_components:
            messagebox.showwarning("Attention", "Aucun composant à exporter")
            return
        
        # Demander le nom du fichier
        filename = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")]
        )
        
        if not filename:
            return
        
        try:
            # Créer le workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM Sélection"
            
            # Styles
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes
            headers = ['Quantité', 'Référence', 'Valeur', 'Footprint', 'LCSC']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Regrouper par valeur et footprint pour compter les quantités
            grouped = {}
            for comp in self.filtered_components:
                key = (comp['value'], comp['footprint'], comp['lcsc'])
                if key not in grouped:
                    grouped[key] = {'refs': [], 'value': comp['value'], 
                                   'footprint': comp['footprint'], 'lcsc': comp['lcsc']}
                grouped[key]['refs'].append(comp['ref'])
            
            # Écrire les données
            row = 2
            for key, data in sorted(grouped.items()):
                refs = sorted(data['refs'])
                ws.cell(row=row, column=1, value=len(refs)).border = thin_border
                ws.cell(row=row, column=2, value=', '.join(refs)).border = thin_border
                ws.cell(row=row, column=3, value=data['value']).border = thin_border
                ws.cell(row=row, column=4, value=data['footprint']).border = thin_border
                ws.cell(row=row, column=5, value=data['lcsc']).border = thin_border
                row += 1
            
            # Ajuster la largeur des colonnes
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            
            # Sauvegarder
            wb.save(filename)
            
            messagebox.showinfo(
                "Succès",
                f"Fichier Excel créé avec succès!\n{filename}"
            )
            self.status_var.set(f"Export réussi: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export:\n{str(e)}")
    
    def _export_csv(self):
        """Exporte les composants sélectionnés vers CSV"""
        if not self.filtered_components:
            messagebox.showwarning("Attention", "Aucun composant à exporter")
            return
        
        # Demander le nom du fichier
        filename = filedialog.asksaveasfilename(
            title="Enregistrer le fichier CSV",
            defaultextension=".csv",
            filetypes=[("Fichiers CSV", "*.csv")]
        )
        
        if not filename:
            return
        
        try:
            # Regrouper par valeur et footprint
            grouped = {}
            for comp in self.filtered_components:
                key = (comp['value'], comp['footprint'], comp['lcsc'])
                if key not in grouped:
                    grouped[key] = {'refs': [], 'value': comp['value'], 
                                   'footprint': comp['footprint'], 'lcsc': comp['lcsc']}
                grouped[key]['refs'].append(comp['ref'])
            
            # Écrire le fichier CSV
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Quantité', 'Référence', 'Valeur', 'Footprint', 'LCSC'])
                
                for key, data in sorted(grouped.items()):
                    refs = sorted(data['refs'])
                    writer.writerow([
                        len(refs),
                        ', '.join(refs),
                        data['value'],
                        data['footprint'],
                        data['lcsc']
                    ])
            
            messagebox.showinfo(
                "Succès",
                f"Fichier CSV créé avec succès!\n{filename}"
            )
            self.status_var.set(f"Export CSV réussi: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export CSV:\n{str(e)}")
    
    # ==================== Gestion de l'historique ====================
    
    def _get_history_file_path(self):
        """Retourne le chemin du fichier d'historique basé sur le fichier HTML chargé"""
        if not self.file_var.get():
            return None
        html_path = Path(self.file_var.get())
        return html_path.parent / f".{html_path.stem}_history.json"
    
    def _load_history(self):
        """Charge l'historique depuis le fichier JSON"""
        self.history = []
        self.current_history_index = None
        self.history_file = self._get_history_file_path()
        
        if self.history_file and self.history_file.exists():
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    self.history = json.load(f)
                print(f"Historique chargé: {len(self.history)} sélections")
            except Exception as e:
                print(f"Erreur lors du chargement de l'historique: {e}")
                self.history = []
        
        self._update_history_combo()
    
    def _save_history(self):
        """Sauvegarde l'historique dans le fichier JSON"""
        if not self.history_file:
            return
        
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, indent=2, ensure_ascii=False)
            print(f"Historique sauvegardé: {len(self.history)} sélections")
        except Exception as e:
            print(f"Erreur lors de la sauvegarde de l'historique: {e}")
    
    def _update_history_combo(self):
        """Met à jour la combobox de l'historique"""
        items = []
        for i, entry in enumerate(self.history):
            name = entry.get('name', f"Sélection {i+1}")
            date = entry.get('date', '')
            count = len(entry.get('components', []))
            processed = len(entry.get('processed', []))
            items.append(f"{name} ({count} comp., {processed} traités) - {date}")
        
        self.history_combo['values'] = items
        if items and self.current_history_index is not None:
            self.history_combo.current(self.current_history_index)
    
    def _on_history_select(self, event=None):
        """Appelé quand on sélectionne un élément dans l'historique"""
        pass  # La sélection est gérée par _load_history_selection
    
    def _load_history_selection(self):
        """Charge la sélection depuis l'historique"""
        if not self.history:
            messagebox.showinfo("Info", "Aucun historique disponible")
            return
        
        selection_idx = self.history_combo.current()
        if selection_idx < 0 or selection_idx >= len(self.history):
            messagebox.showwarning("Attention", "Veuillez sélectionner une entrée dans l'historique")
            return
        
        entry = self.history[selection_idx]
        self.current_history_index = selection_idx
        
        # Restaurer la zone de sélection
        rect = entry.get('rect')
        if rect and len(rect) == 4:
            self.selection_rect = tuple(rect)
            # Récupérer les composants dans cette zone
            self.selected_components = self.parser.get_components_in_rect(*self.selection_rect)
        else:
            # Utiliser les composants sauvegardés directement
            saved_refs = set(c.get('ref') for c in entry.get('components', []))
            self.selected_components = [
                comp for comp in self.parser.components
                if comp.get('ref') in saved_refs
            ]
            # Ajouter les infos BOM
            for comp in self.selected_components:
                bom_info = self.parser.get_bom_for_ref(comp['ref'], comp.get('id'))
                comp['value'] = bom_info.get('value', '')
                comp['footprint'] = bom_info.get('footprint', '')
                comp['lcsc'] = bom_info.get('lcsc', '')
        
        # Restaurer les items traités
        self.processed_items.clear()
        for proc in entry.get('processed', []):
            if isinstance(proc, list) and len(proc) == 3:
                self.processed_items.add(tuple(proc))
        
        self._apply_filters()
        self._draw_pcb_preview()
        
        self.export_btn.config(state=tk.NORMAL)
        self.export_csv_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)
        
        name = entry.get('name', f"Sélection {selection_idx + 1}")
        self.status_var.set(f"Chargé: {name} ({len(self.selected_components)} composants)")
    
    def _save_current_to_history(self):
        """Sauvegarde la sélection actuelle dans l'historique"""
        if not self.selected_components:
            messagebox.showwarning("Attention", "Aucune sélection à sauvegarder")
            return
        
        # Demander un nom pour la sélection
        from tkinter import simpledialog
        name = simpledialog.askstring(
            "Nom de la sélection",
            "Entrez un nom pour cette sélection:",
            initialvalue=f"Zone {len(self.history) + 1}"
        )
        
        if not name:
            return
        
        # Créer l'entrée d'historique
        entry = {
            'name': name,
            'date': datetime.now().strftime("%Y-%m-%d %H:%M"),
            'rect': list(self.selection_rect) if self.selection_rect else None,
            'components': [
                {'ref': c['ref'], 'value': c['value'], 'footprint': c['footprint'], 'lcsc': c['lcsc']}
                for c in self.selected_components
            ],
            'processed': [list(p) for p in self.processed_items]
        }
        
        self.history.append(entry)
        self.current_history_index = len(self.history) - 1
        self._save_history()
        self._update_history_combo()
        
        messagebox.showinfo("Succès", f"Sélection '{name}' sauvegardée dans l'historique")
    
    def _update_history_selection(self):
        """Met à jour l'entrée d'historique actuelle avec les modifications"""
        if not self.history:
            messagebox.showinfo("Info", "Aucun historique disponible")
            return
        
        selection_idx = self.history_combo.current()
        if selection_idx < 0 or selection_idx >= len(self.history):
            messagebox.showwarning("Attention", "Veuillez sélectionner une entrée dans l'historique")
            return
        
        if not self.selected_components:
            messagebox.showwarning("Attention", "Aucune sélection active à mettre à jour")
            return
        
        # Mettre à jour l'entrée
        entry = self.history[selection_idx]
        entry['date'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        entry['rect'] = list(self.selection_rect) if self.selection_rect else None
        entry['components'] = [
            {'ref': c['ref'], 'value': c['value'], 'footprint': c['footprint'], 'lcsc': c['lcsc']}
            for c in self.selected_components
        ]
        entry['processed'] = [list(p) for p in self.processed_items]
        
        self._save_history()
        self._update_history_combo()
        
        messagebox.showinfo("Succès", f"Sélection '{entry['name']}' mise à jour")
    
    def _delete_history_selection(self):
        """Supprime l'entrée d'historique sélectionnée"""
        if not self.history:
            messagebox.showinfo("Info", "Aucun historique disponible")
            return
        
        selection_idx = self.history_combo.current()
        if selection_idx < 0 or selection_idx >= len(self.history):
            messagebox.showwarning("Attention", "Veuillez sélectionner une entrée dans l'historique")
            return
        
        entry = self.history[selection_idx]
        name = entry.get('name', f"Sélection {selection_idx + 1}")
        
        if messagebox.askyesno("Confirmation", f"Supprimer la sélection '{name}' de l'historique?"):
            del self.history[selection_idx]
            self.current_history_index = None
            self._save_history()
            self._update_history_combo()
            self.history_var.set("")
            messagebox.showinfo("Succès", f"Sélection '{name}' supprimée")
    
    def _setup_keyboard_shortcuts(self):
        """Configure les raccourcis clavier"""
        self.root.bind('<Control-o>', lambda e: self._browse_file())
        self.root.bind('<Control-s>', lambda e: self._export_excel() if self.filtered_components else messagebox.showinfo("Info", "Aucun composant sélectionné"))
        self.root.bind('<Control-Shift-s>', lambda e: self._export_csv() if self.filtered_components else messagebox.showinfo("Info", "Aucun composant sélectionné"))
        self.root.bind('<Control-l>', lambda e: self._load_file())
        self.root.bind('<Escape>', lambda e: self._clear_selection())
        self.root.bind('<Control-f>', lambda e: self.search_entry.focus_set() if hasattr(self, 'search_entry') else None)
        self.root.bind('<F5>', lambda e: self._draw_pcb_preview() if self.parser else None)
    
    def run(self):
        """Lance l'application"""
        # Charger automatiquement bom/ibom.html s'il existe
        self._auto_load_bom()
        self.root.mainloop()
    
    def _auto_load_bom(self):
        """Charge automatiquement un fichier ibom.html du répertoire bom si présent"""
        import os
        
        # Chercher dans le répertoire bom relatif au script
        script_dir = Path(__file__).parent
        bom_paths = [
            script_dir / 'bom' / 'ibom.html',
            script_dir / 'bom' / 'bom.html',
            Path('bom') / 'ibom.html',
            Path('bom') / 'bom.html',
        ]
        
        for bom_path in bom_paths:
            if bom_path.exists():
                self.file_var.set(str(bom_path.resolve()))
                self.root.after(100, self._load_file)  # Charger après l'affichage
                break


if __name__ == '__main__':
    app = IBomSelectorApp()
    app.run()
